import glob
import itertools
import os
import re
import time
from collections import OrderedDict, defaultdict
from multiprocessing import Pool, cpu_count

import SimpleITK as sitk
import cv2
import matplotlib.pyplot as plt
import numpy as np
import pandas as pd
import pydicom
import scipy.ndimage as ndi
from openpyxl import load_workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
from radiomics import featureextractor
from scipy import signal
from scipy import stats
from scipy.fft import fft
from scipy.integrate import simps
from scipy.interpolate import interp1d
from scipy.ndimage import binary_dilation
from scipy.ndimage import binary_erosion
from scipy.stats import mannwhitneyu
from scipy.stats import skew, kurtosis
from skimage.measure import shannon_entropy
from skimage.metrics import peak_signal_noise_ratio, structural_similarity, mean_squared_error
from skimage.segmentation import find_boundaries
from skimage.transform import resize
from sklearn.metrics import roc_auc_score


def histogram_asymmetry_ratio(img: np.ndarray, mask: np.ndarray, use_median=False):
    """
    计算正负像素强度面积比（对称性指标）
    - 如果完全对称，则结果接近1
    - 偏左或偏右 => 结果远离1

    参数：
        img : 图像
        mask : 掩膜
        use_median : 是否使用中位数作为分界（默认使用均值）

    返回：
        asymmetry_ratio : 正 / 负 强度面积比
    """
    pixels = img[mask > 0].astype(np.float32)
    # print(pixels)
    if pixels.size == 0:
        return np.nan

    center_val = np.median(pixels) if use_median else np.mean(pixels)
    pos_sum = np.sum(pixels[pixels > center_val] - center_val)
    neg_sum = np.sum(center_val - pixels[pixels < center_val])

    if neg_sum == 0:
        return np.inf  # 防止除以0
    return pos_sum / neg_sum


def count_ones_per_layer(mask_3d):
    if mask_3d.ndim != 3:
        raise ValueError("输入的 mask_3d 必须是三维数组")
    print("\n".join(f"第 {i} 层 1 的个数: {count}"
                    for i, count in enumerate(np.sum(mask_3d == 1, axis=(1, 2)))))


def get_top_z_indices_joint_mask(mask_3d, mask_3d_control, top_n=5):
    """
    获取两个掩膜都存在的层中，结构掩膜值最大的 top_n 层索引
    """
    # 获取结构掩膜中每层的 1 的个数
    ones_counts = np.sum(mask_3d == 1, axis=(1, 2))

    # 获取同时存在结构结构和控制结构的层索引
    valid_indices = [z for z in range(mask_3d.shape[0])
                     if np.any(mask_3d[z]) and np.any(mask_3d_control[z])]

    # 针对这些有效层，提取其 ones_counts
    valid_counts = [(z, ones_counts[z]) for z in valid_indices]

    # 根据 ones_counts 倒序排序
    sorted_valid = sorted(valid_counts, key=lambda x: x[1], reverse=True)

    # 取前 top_n 层的索引
    top_z_indices = [z for z, _ in sorted_valid[:top_n]]

    # 如果selected_metric存在 则改变计算层面为以这个指标最大值来计算层面 具体方法在后面计算完后的时候改

    return top_z_indices


def change_count_ones_per_layer(mask_3d, th=90):
    """
    计算并打印 mask_3d 在每一层中 1 的个数，并将 1 个数小于 80 的层设为 0。

    参数：
    mask_3d (numpy.ndarray): 三维掩膜数组
    """
    if not isinstance(mask_3d, np.ndarray) or mask_3d.ndim != 3:
        raise ValueError("输入的 mask_3d 必须是三维 NumPy 数组")
    ones_counts = np.sum(mask_3d == 1, axis=(1, 2))
    mask_3d[ones_counts < th] = 0  # 将 1 个数小于 80 的层设为 0


def analyze_ct_values(ori_data_output, analyze_data_output):
    # 读取Excel数据
    df = pd.read_excel(ori_data_output)

    # 获取输出目录路径，确保文件夹存在
    output_dir = os.path.dirname(analyze_data_output)
    if not os.path.exists(output_dir):
        os.makedirs(output_dir)

    results = []
    folders = df['Folder'].unique()
    Metrics = df['Metric'].unique()

    # 遍历不同 Folder 下相同 Metric 的 CT 值
    for Metric in Metrics:
        print(f"Comparing for Metric: {Metric}")
        for i in range(len(folders)):
            for j in range(i + 1, len(folders)):
                folder1 = folders[i]
                folder2 = folders[j]

                # 获取数据
                ct_values_folder1 = df[(df['Folder'] == folder1) & (df['Metric'] == Metric)]['Value'].values
                ct_values_folder2 = df[(df['Folder'] == folder2) & (df['Metric'] == Metric)]['Value'].values

                # 处理数据长度不匹配的情况
                if len(ct_values_folder1) != len(ct_values_folder2):
                    print(f'统计检验   {folder1} vs {folder2}   指标： ({Metric})    CT 值长度不匹配: {len(ct_values_folder1)} vs {len(ct_values_folder2)}  进行裁剪')

                    diff_len = abs(len(ct_values_folder1) - len(ct_values_folder2))
                    if len(ct_values_folder1) > len(ct_values_folder2):
                        ct_values_folder1 = np.partition(ct_values_folder1, diff_len)[:-diff_len]
                    else:
                        ct_values_folder2 = np.partition(ct_values_folder2, diff_len)[:-diff_len]

                    # print(f'   处理后长度: {len(ct_values_folder1)} vs {len(ct_values_folder2)}')

                # 计算均值
                avg1 = np.mean(ct_values_folder1)
                avg2 = np.mean(ct_values_folder2)

                # **正态性检验（Shapiro-Wilk）**
                try:
                    p_shapiro1 = stats.shapiro(ct_values_folder1).pvalue if len(ct_values_folder1) <= 50 else np.nan
                    p_shapiro2 = stats.shapiro(ct_values_folder2).pvalue if len(ct_values_folder2) <= 50 else np.nan
                except ValueError:
                    p_shapiro1, p_shapiro2 = np.nan, np.nan

                normality_folder1 = 'Yes' if p_shapiro1 > 0.05 else 'No'
                normality_folder2 = 'Yes' if p_shapiro2 > 0.05 else 'No'

                # **配对 t 检验**
                try:
                    t_stat, p_ttest = stats.ttest_rel(ct_values_folder1, ct_values_folder2)
                except ValueError:
                    t_stat, p_ttest = np.nan, np.nan

                # **Wilcoxon 符号秩检验**
                try:
                    w_stat, p_wilcoxon = stats.wilcoxon(ct_values_folder1, ct_values_folder2)
                except ValueError:
                    w_stat, p_wilcoxon = np.nan, np.nan

                # **存储所有计算结果**
                results.append({
                    'ROI': Metric,
                    'Folder1': folder1,
                    'Folder2': folder2,
                    'Mean1': avg1,
                    'Mean2': avg2,
                    'ΔMean': avg1 - avg2,
                    'Shapiro P1': p_shapiro1,
                    'Shapiro P2': p_shapiro2,
                    'Normality1': normality_folder1,
                    'Normality2': normality_folder2,
                    'T-test Statistic': t_stat,
                    'T-test P-value': p_ttest,
                    'T-test Significant': 'Yes' if p_ttest < 0.05 else 'No',
                    'Wilcoxon Statistic': w_stat,
                    'Wilcoxon P-value': p_wilcoxon,
                    'Wilcoxon Significant': 'Yes' if p_wilcoxon < 0.05 else 'No'
                })

    # **保存结果**
    results_df = pd.DataFrame(results)
    results_df.to_csv(analyze_data_output, index=False)
    print(f"统计检验结果保存到 {analyze_data_output}")


def load_dicom_images(ct_folder_path):
    """
    加载指定文件夹中的所有 DICOM CT 切片，并按 Z 轴(ImagePositionPatient[2]) 排序。
    将像素值转换到 (约为) HU 空间后，组合成一个 3D 体数据 volume。

    返回:
        volume (numpy.ndarray): 3D CT图像, shape = (num_slices, height, width), int16, 已做HU校正(若Slope/Intercept存在).
        origin (numpy.ndarray): 物理坐标系原点, [x0, y0, z0].
        spacing (numpy.ndarray): 像素间距 [row_spacing, col_spacing, slice_thickness].
        z_max, y_max, x_max (int): volume各维度大小.
        dcm_slices (list of pydicom.dataset.FileDataset): 按Z排序后的 DICOM 对象列表(与 volume 对应).
    """
    # 1) 找到所有 DICOM 文件
    dicom_paths = glob.glob(os.path.join(ct_folder_path, "CT*.dcm"))
    if not dicom_paths:
        raise ValueError(f"在文件夹 {ct_folder_path} 中未找到任何 .dcm 文件！")

    # 2) 读取并筛选含有效 ImagePositionPatient 的切片
    dcm_slices = []
    for path in dicom_paths:
        ds = pydicom.dcmread(path)
        if hasattr(ds, 'ImagePositionPatient'):
            dcm_slices.append(ds)
        else:
            print(f"文件 {path} 缺少 ImagePositionPatient 属性，已跳过。")
    if not dcm_slices:
        raise ValueError("没有找到包含 ImagePositionPatient 的 DICOM 切片，无法重建。")

    # 3) 按 Z 坐标排序
    # dcm_slices.sort(key=lambda ds: float(ds.ImagePositionPatient[2]))
    dcm_slices.sort(key=lambda ds: ds.InstanceNumber)
    # 4) 获取基本元数据 (取第一张切片为代表)
    first_slice = dcm_slices[0]
    # 像素间距(行、列)
    row_spacing, col_spacing = map(float, first_slice.PixelSpacing)
    slice_thickness = float(first_slice.SliceThickness)
    # slice_thickness = 1.5
    # print('注意之前CT成像参数填错了实际slice_thickness = 1.5， 目前是强制改成1.5')
    # 原点(仅供参考：DICOM 的 (x,y,z), 这里取第一片的 ImagePositionPatient)
    origin = np.array(first_slice.ImagePositionPatient, dtype=float)
    # date
    date_day = first_slice.AcquisitionDate
    acquisitionTime = first_slice.AcquisitionTime
    date = f'{date_day}' + f'-{acquisitionTime}'
    # print(acquisitionTime)
    # print(date)
    # ID
    ID = first_slice.PatientID
    # 构建 spacing=[row_spacing, col_spacing, slice_thickness]
    # 注意: DICOM 中 PixelSpacing 顺序一般是 [row_spacing, col_spacing],
    #       也常常对应 [Y_spacing, X_spacing], 具体要看 orientation
    spacing = np.array([row_spacing, col_spacing, slice_thickness], dtype=float)

    # 5) 建立 volume 数组 (num_slices, height, width)
    num_slices = len(dcm_slices)
    height = first_slice.pixel_array.shape[0]
    width = first_slice.pixel_array.shape[1]
    volume = np.zeros((num_slices, height, width), dtype=np.int16)

    # 6) 将每张切片读入 volume
    for i, ds in enumerate(dcm_slices):
        arr = ds.pixel_array.astype(np.int16)  # 原始灰度
        volume[i, :, :] = arr

    # 7) 进行 HU 转换 (使用 RescaleSlope/Intercept)
    for i, ds in enumerate(dcm_slices):
        slope = ds.RescaleSlope if 'RescaleSlope' in ds else 1.0
        intercept = ds.RescaleIntercept if 'RescaleIntercept' in ds else 0.0

        arr_float = volume[i].astype(np.float32)
        if slope != 1.0:
            arr_float *= slope
        arr_float += intercept
        volume[i] = arr_float.astype(np.int16)

    z_max, y_max, x_max = volume.shape
    # print(f"             Loaded volume shape = {z_max} x {y_max} x {x_max}")
    # print(f"             Origin = {origin}, Spacing = {spacing}")
    # print('             图像HU值缩放系数', intercept, slope)
    return volume, origin, spacing, z_max, y_max, x_max, dcm_slices, date_day, date, ID


def mtf_calcu(
        image):  # 抄的来自于CBCTSPC项目中的自己编写的简单MTF，建议一起算个结果万一有用
    """
    :param image:
    :return:
    """
    mtf_areas_row = []  # 用于存储每一行的MTF曲线下面积
    mtf_areas_col = []
    # 遍历图像的每一行
    for row in image:
        # 计算ESF：直接使用该行的像素值
        esf = row
        if esf.size < 2 or np.allclose(esf, esf[0]):
            continue  # 跳过无效行
        # 计算LSF：ESF的一阶导数
        lsf = np.gradient(esf)
        # 计算MTF：对LSF进行傅里叶变换并归一化
        mtf = np.abs(fft(lsf))
        max_mtf = np.max(mtf)
        if max_mtf == 0 or np.isnan(max_mtf):
            continue  # 跳过异常行，避免除以0
        mtf = mtf / max_mtf
        # 计算MTF曲线下面积，使用辛普森规则
        area = simps(mtf, dx=1)
        mtf_areas_row.append(area)  # 反映的是竖直边缘的mtf
    # 返回所有MTF曲线下面积的平均值
    for col_idx in range(image.shape[1]):  # 遍历列索引
        col = image[:, col_idx]  # 提取该列
        # 计算ESF：直接使用该行的像素值
        esf = col
        if esf.size < 2 or np.allclose(esf, esf[0]):
            continue  # 跳过无效列
        # 计算LSF：ESF的一阶导数
        lsf = np.gradient(esf)
        # 计算MTF：对LSF进行傅里叶变换并归一化
        mtf = np.abs(fft(lsf))
        max_mtf = np.max(mtf)
        if max_mtf == 0 or np.isnan(max_mtf):
            continue  # 跳过异常行，避免除以0
        mtf = mtf / max_mtf
        # 计算MTF曲线下面积，使用辛普森规则
        area = simps(mtf, dx=1)
        mtf_areas_col.append(area)
    # 返回所有MTF曲线下面积的平均值
    return np.mean(mtf_areas_row), np.mean(mtf_areas_col)


# 计算每个切片的 MTF 面积（积分）
def get_mtf_area(mtf_dict):
    mtf_areas = {}
    for slice_idx, (freqs, avg_mtf) in mtf_dict.items():
        area = simps(avg_mtf, freqs)
        mtf_areas[slice_idx] = area
    return mtf_areas


# 计算每个切片的 MTF 截止频率
def get_mtf_cutoff_freq(mtf_dict, threshold=0.1):
    cutoff_freqs = {}
    for slice_idx, (freqs, avg_mtf) in mtf_dict.items():
        cutoff_freq = np.interp(threshold, avg_mtf[::-1], freqs[::-1])  # 反向查找 10% MTF 对应的频率
        cutoff_freqs[slice_idx] = cutoff_freq
    return cutoff_freqs


def compute_single_profile_mtf(roi, ID, date, z, profile, Fs, smoothing_window=None, interp_points=100):
    """
    计算单条剖面的 MTF 曲线、面积和 10% 截止频率（支持剖面插值）
    参数:
        profile: 原始 1D 边缘剖面 (numpy array)
        Nfft: FFT 点数
        Fs: 采样频率 (pixel_size_mm^-1)
        smoothing_window: 平滑窗口（可选）
        interp_factor: 插值100数据点  统一
    """
    # ---------- 0. 健壮性检查 ----------
    if profile.size < 2 or np.allclose(profile, profile[0]):
        # print(profile)
        # raise ValueError(f"{roi,ID,date,z,}profile must have at least 2 points and not be constant.")
        return 0, 0  # 或 (None, None) 看你后续如何处理
    # 1. 插值 profile 到 interp_points 长度
    x_old = np.linspace(0, 1, len(profile))
    x_new = np.linspace(0, 1, interp_points)
    # bounds_error=False 可以避免超界直接报错
    profile = interp1d(
        x_old, profile,
        kind='linear',
        bounds_error=False,
        fill_value=(profile[0], profile[-1])
    )(x_new)
    esf = profile.astype(np.float32)
    esf -= esf.min()
    peak = esf.max()
    if peak <= 1e-9:
        return 0, 0
    esf /= peak
    lsf = np.gradient(esf)
    Nfft = 1 << (interp_points - 1).bit_length()  # 自动确定下一个2的幂
    LSF_fft = fft(lsf, n=Nfft)
    mag = np.abs(LSF_fft[:Nfft // 2])
    if mag.max() == 0:
        return 0, 0
    mag /= mag.max()
    freqs = np.linspace(0, Fs / 2, Nfft // 2, endpoint=False)
    if smoothing_window is not None:
        mag = signal.savgol_filter(mag, smoothing_window, 3)
    mtf_area = simps(mag, freqs)
    try:
        cutoff = np.interp(0.1, mag[::-1], freqs[::-1])
    except Exception:
        cutoff = np.nan
    return mtf_area, cutoff


def compute_mtf_both_axes(roi, ID, date, z,
                          image_2d: np.ndarray,
                          pixel_size_mm: float = 1.0,
                          smoothing_window: int = 11
                          ):
    # assert isinstance(image_2d, np.ndarray), f"metric_mtf: 'img' {roi}{ID}{date}{z}不是 ndarray，而是 {type(image_2d)}"
    # print()
    rows, cols = image_2d.shape
    Fs = 1.0 / pixel_size_mm
    mtf_areas_row, cutoff_freqs_row = [], []
    for r in range(rows):
        profile = image_2d[r, :]
        area, cutoff = compute_single_profile_mtf(roi, ID, date, z, profile, Fs)
        if area is not None:
            mtf_areas_row.append(area)
            cutoff_freqs_row.append(cutoff)
    mtf_areas_col, cutoff_freqs_col = [], []
    for c in range(cols):
        profile = image_2d[:, c]
        area, cutoff = compute_single_profile_mtf(roi, ID, date, z, profile, Fs)
        if area is not None:
            mtf_areas_col.append(area)
            cutoff_freqs_col.append(cutoff)
    return np.mean(mtf_areas_row), np.mean(cutoff_freqs_row), np.mean(mtf_areas_col), np.mean(cutoff_freqs_col)


def load_structure_mask(rs_file, volume_shape, dcm_slices, structure_name="A"):
    """
    根据 RS 文件, 生成指定结构(如 "A")在 3D volume (z,y,x) 上的掩膜 (0/1).
    简化假设: x->列, y->行, z->层, orientation=[1,0,0;0,1,0].

    参数:
        rs_file: RT 结构文件路径 (字符串).
        volume_shape: (num_slices, height, width) 与 volume 一致.
        dcm_slices: 已排好序的切片列表(与 volume 对应).
        structure_name: 结构名, 如 "A".

    返回:
        mask_3d: shape 同 volume, dtype=uint8, 0或1.
    """
    rs = pydicom.dcmread(rs_file)
    num_slices, height, width = volume_shape
    mask_3d = np.zeros((num_slices, height, width), dtype=np.uint8)

    # 1) 找到指定结构 ROINumber
    target_roi_num = None
    for roi in rs.StructureSetROISequence:
        if roi.ROIName.lower() == structure_name.lower():
            target_roi_num = roi.ROINumber
            break
    if target_roi_num is None:
        # raise ValueError(f"结构 '{structure_name}' 在 RS 文件中未找到")
        return mask_3d

    # 2) ROIContourSequence 中找到该 ROINumber
    target_roi_contour = None
    for c in rs.ROIContourSequence:
        if c.ReferencedROINumber == target_roi_num:
            target_roi_contour = c
            break
    if not target_roi_contour:
        # raise ValueError(f"ROIContour for '{structure_name}' not found in RS!")
        return mask_3d
    # 3) 遍历 contour
    for contour_seq in target_roi_contour.ContourSequence:
        contour_data = contour_seq.ContourData  # [x1,y1,z1, x2,y2,z2, ...]
        coords = np.array(contour_data).reshape(-1, 3)
        # 找到匹配的切片
        ref_sop = contour_seq.ContourImageSequence[0].ReferencedSOPInstanceUID
        slice_idx = None
        for i, ds in enumerate(dcm_slices):
            if ds.SOPInstanceUID == ref_sop:
                slice_idx = i
                break
        if slice_idx is None:
            continue

        ds_ref = dcm_slices[slice_idx]
        origin_xy = np.array(ds_ref.ImagePositionPatient[:2], dtype=float)
        spacing_xy = np.array(ds_ref.PixelSpacing, dtype=float)  # [row_spacing, col_spacing]

        # 这里简化: row = (Y - origin_y) / row_spacing; col = (X - origin_x) / col_spacing
        poly_points = []
        for X, Y, Z in coords:
            row = (Y - origin_xy[1]) / spacing_xy[0]
            col = (X - origin_xy[0]) / spacing_xy[1]
            poly_points.append([row, col])
        poly_points = np.round(poly_points).astype(np.int32)

        mask_slice = np.zeros((height, width), dtype=np.uint8)
        # OpenCV 填充多边形时, 需要 [ [col,row], ... ], 但 poly_points 是 [row,col],
        # 所以要翻转下 shape -> (N,2)
        cv2.fillPoly(mask_slice, [poly_points[:, [1, 0]]], 1)

        mask_3d[slice_idx] = np.logical_or(mask_3d[slice_idx], mask_slice).astype(np.uint8)

    return mask_3d


def summarize_mask_dict_structure(d, max_depth=4):
    """
    打印每一层的 key 数量，并统计最底层（叶子节点）数量
    """
    from collections import defaultdict

    depth_counts = defaultdict(int)
    leaf_count = 0

    def traverse(node, depth):
        nonlocal leaf_count
        if depth > max_depth:
            return
        if isinstance(node, dict):
            depth_counts[depth] += len(node)
            for v in node.values():
                traverse(v, depth + 1)
        else:
            leaf_count += 1

    traverse(d, depth=0)

    print("        📊 各层字典结构:")
    for depth in sorted(depth_counts):
        print(f"          第 {depth} 层: {depth_counts[depth]} 个键")
    print(f"        ✅ 最小单元（叶子节点）总数: {leaf_count}")


def plt_image2(img, z, ID, date, window_width=400, window_level=40):
    # 确保窗宽窗位应用到图像的范围
    img_min = window_level - window_width / 2
    img_max = window_level + window_width / 2

    # 对图像进行窗宽窗位的线性变换
    img_display = np.clip(img, img_min, img_max)  # 限制像素值在窗宽窗位范围内
    img_display = (img_display - img_min) / (img_max - img_min) * 255  # 线性映射到[0, 255]

    # 使用plotly显示图像
    fig = px.imshow(np.array(img_display), color_continuous_scale='gray', color_continuous_midpoint=(0 + 255) / 2)
    # 设置固定的色阶范围
    fig.update_layout(
        title=f"ID{ID}   date{date}   Slice {z}",
        coloraxis=dict(
            colorbar=dict(title='Intensity'),
            colorscale='gray',
            cmin=0,  # 固定最小灰度值
            cmax=255  # 固定最大灰度值
        )
    )
    fig.show()


def plt_image_with_mask_and_save(img, mask, tag='额外标签', z=None, ID=None, date=None):
    # 窗宽窗位调整
    img_min = -125
    img_max = 225
    img_display = np.clip(img, img_min, img_max)
    img_display = ((img_display - img_min) / (img_max - img_min) * 255).astype(np.uint8)
    # 创建 RGB 图像
    img_rgb = np.stack([img_display] * 3, axis=-1)  # 转为 3 通道
    # 把 mask 区域设为红色
    red_mask = mask > 0
    img_rgb[red_mask] = [255, 255, 255]  # 红色完全覆盖
    # 显示图像
    plt.figure(figsize=(6, 6))
    plt.imshow(img_rgb, interpolation='none')
    plt.title(f"Slice {z} | ID: {ID} | Date: {date} | tag: {tag}")
    plt.axis('off')
    # plt.tight_layout()
    plt.show()
    plt.savefig(os.path.join(r"D:\图片", f"{ID}_{date}_{tag}.png"))


def plt_binary_mask(mask):
    plt.figure(figsize=(5, 5))
    plt.imshow(mask, cmap='gray')  # 使用灰度颜色映射
    plt.title(f"Slice  - Binary Mask")
    plt.axis('off')  # 不显示坐标轴
    plt.tight_layout()
    plt.show()


def calculate_nps(image, spacing, tag=''):
    # plt_image2(image,2)
    # 1. 确保输入图像是二维HU矩阵
    if image.ndim != 2:
        raise ValueError("输入图像必须是二维矩阵")
    # 2. 对图像进行傅里叶变换
    image = image - np.mean(image)  # 去除直流分量
    fft_image = np.fft.fftshift(np.fft.fft2(image))
    # 3. 计算功率谱   # 原始二维的NPS
    power_spectrum = np.abs(fft_image) ** 2
    # 4. 计算频率的二维坐标系，中心化
    dx, dy = spacing[1], spacing[0]  # 从 CT 图像的 spacing 参数获取像素间距 目的： 乘以 采样率（像素间距的倒数），以转换为实际的 空间频率（单位：mm⁻¹）。
    freqs_x = np.fft.fftfreq(image.shape[1], d=dx)  # X方向空间频率 (mm⁻¹)
    freqs_y = np.fft.fftfreq(image.shape[0], d=dy)  # Y方向空间频率 (mm⁻¹)
    freqs_x, freqs_y = np.meshgrid(freqs_x, freqs_y)
    # 5. 可选：计算环形平均功率谱
    r = np.sqrt(freqs_x ** 2 + freqs_y ** 2)
    r = np.floor(r * max(image.shape))  # 量化为整数索引
    # 计算环形平均
    bins = np.arange(0, np.max(r) + 1)
    nps, _ = np.histogram(r, bins=bins, weights=power_spectrum)
    counts, _ = np.histogram(r, bins=bins)
    nps = nps / counts
    # 6. 计算每个频率点的角度（以度为单位）
    angles = np.arctan2(freqs_y, freqs_x) * 180 / np.pi
    # 7. 计算方向性功率谱  中心化后 0度为x-正方向,计算每个1度范围的方向性噪声  并按照金属伪影的方向 根据不同的mask位置来计算方向性噪声
    angle_range = 1  # 每1度范围
    # # 根据不同的方向进行计算
    if tag == 'up':
        directional_nps = calculate_directional_power(angles, power_spectrum, 0, 180, angle_range)
    elif tag == 'down':
        directional_nps = calculate_directional_power(angles, power_spectrum, -180, 0, angle_range)
    elif tag == 'left':
        directional_nps = calculate_directional_power(angles, power_spectrum, 90, 270, angle_range)
    elif tag == 'right':
        directional_nps = calculate_directional_power(angles, power_spectrum, -90, 90, angle_range)
    # elif tag == 'all':  # 否则不计算该值 默认1
    #     # directional_nps = [1,1]
    #     directional_nps = calculate_directional_power(angles, power_spectrum, 0, 360, angle_range)
    else:  # 否则计算360度的
        directional_nps = calculate_directional_power(angles, power_spectrum, 0, 360, angle_range)
        # print(f'当前tag不太对  转换为计算all:{tag}')
        # raise ValueError('计算NPS方向参数不对')
    # 8. 设定频率阈值划分高低频
    f_threshold = np.max(r) / 2  # 设定低频和高频的分界线
    low_freq_mask = r < f_threshold  # 低频部分的掩码
    high_freq_mask = r >= f_threshold  # 高频部分的掩码
    # 计算低频和高频噪声能量
    low_freq_nps = np.sum(power_spectrum[low_freq_mask])
    high_freq_nps = np.sum(power_spectrum[high_freq_mask])
    # 9. 计算能量总和：积分求和
    total_power_spectrum = np.sum(power_spectrum)  # 计算总功率谱的能量
    total_nps_ring = np.sum(nps)  # 计算环形平均功率谱的能量
    total_directional_power = np.sum(directional_nps)  # 计算方向性功率谱的能量
    return total_power_spectrum, total_nps_ring, total_directional_power, directional_nps, low_freq_nps, high_freq_nps


def calculate_directional_power(angles, power_spectrum, start_angle, end_angle, angle_range):
    directional_power = np.zeros((end_angle - start_angle) // angle_range)
    for i in range(directional_power.shape[0]):
        lower_bound = start_angle + i * angle_range
        upper_bound = start_angle + (i + 1) * angle_range
        direction_mask = (angles >= lower_bound) & (angles < upper_bound)
        directional_power[i] = np.sum(power_spectrum[direction_mask])
    return directional_power


def create_remaining_masks(folder_name, base_mask, pixels_far, mask_tags=None, para_name=None):
    """
       根据 mask_tags 生成掩膜集合。
       - base_mask : numpy array，原 ROI mask（0/1）
       - pixels_far: 与 base_mask 保持的最小距离
       - mask_tags : 想要的子掩膜标签顺序，可扩展 ('tag1', 'tag2', ...)
       返回 OrderedDict: {tag1: mask1, tag2: mask2, ...}
       """
    # 获取矩形区域的坐标
    ys, xs = np.where(base_mask > 0)
    y_min, y_max = ys.min(), ys.max()
    x_min, x_max = xs.min(), xs.max()
    # H, W = base_mask.shape
    masks = OrderedDict()
    for tag in mask_tags:
        m = np.zeros_like(base_mask)
        # if tag in ['up', 'down', 'left', 'right'] and para_name != 'spc1_for_MAS':
        #     y_start = np.clip(y_min - pixels_far, 0, H)
        #     x_start = np.clip(x_min - pixels_far, 0, W)
        #     y_end= np.clip(y_max + pixels_far, 0, H)
        #     x_end = np.clip(x_max + pixels_far, 0, W)
        #     if tag == 'up':
        #         m[:y_start,x_start:x_end] = 1
        #     elif tag == 'down':
        #         m[y_end:,x_start:x_end] = 1
        #     elif tag == 'left':
        #         m[y_start:y_end,:x_start] = 1
        #     elif tag == 'right':
        #         m[y_start:y_end,x_end:] = 1
        # elif tag in ['up', 'down', 'left', 'right'] and para_name == 'spc1_for_MAS':
        #     if tag == 'up':
        #         m[:y_min - pixels_far, x_min - pixels_far:x_max + pixels_far] = 1
        #     elif tag == 'down':
        #         m[y_max + pixels_far:, x_min - pixels_far:x_max + pixels_far] = 1
        #     elif tag == 'left':
        #         m[:y_max + pixels_far, :x_min - pixels_far] = 1
        #     elif tag == 'right':
        #         m[:y_max + pixels_far, x_max + pixels_far:] = 1
        if tag in ['up', 'down', 'left', 'right']:  # 不考虑越界裁剪 全部都统一生成
            if tag == 'up':
                m[:y_min - pixels_far, x_min - pixels_far:x_max + pixels_far] = 1
            elif tag == 'down':
                m[y_max + pixels_far:, x_min - pixels_far:x_max + pixels_far] = 1
            elif tag == 'left':
                m[:y_max + pixels_far, :x_min - pixels_far] = 1
                # if folder_name == '2024-09-24-006_20241231_jian':
                #     H, W = m.shape
                #     print(f" mask 尺寸: 高度 H = {H}（y: 0 ~ {H - 1}），宽度 W = {W}（x: 0 ~ {W - 1}）")
                #     print('mask范围参数', y_min, y_max, x_min, x_max, pixels_far)
                #     print(f'左部: 0-{y_max + pixels_far}行    0-{x_min - pixels_far} 列')
                #     print(f'右部: 0-{y_max + pixels_far}行    {x_max + pixels_far}-{W}列')
            elif tag == 'right':
                m[:y_max + pixels_far, x_max + pixels_far:] = 1
        # 对应rectum_dilated_upper_half 原本用于计算全参考指标 因此只需要扩大一下subimg即可 不需要派生什么特殊的子roi
        elif tag == 'rectum_dilated_upper_half':
            m = base_mask
        else:
            raise ValueError(f'未知 mask_tag: {tag}')
        masks[tag] = m.astype(np.uint8)

    return masks


def create_remaining_masks2(mask, pixels_far):
    # 获取矩形区域的坐标
    ys, xs = np.where(mask > 0)
    y_min, y_max = ys.min(), ys.max()
    x_min, x_max = xs.min(), xs.max()
    # 创建四个新掩膜，初始化为零
    mask1 = np.zeros_like(mask)  # 上部矩形区域
    mask2 = np.zeros_like(mask)  # 下部矩形区域
    mask3 = np.zeros_like(mask)  # 左部矩形区域
    mask4 = np.zeros_like(mask)  # 右部矩形区域
    # 上部矩形区域  再远离结构2像素  旧版y还继续远离2像素  好像没必要远离这么远
    mask1[:y_min - pixels_far, x_min - pixels_far:x_max + pixels_far] = 1  # 从上到M区域上方
    # 下部矩形区域
    # print(y_max+pixels_far, x_min-pixels_far,x_max+pixels_far)
    mask2[y_max + pixels_far:, x_min - pixels_far:x_max + pixels_far] = 1  # 从M区域下方到底部
    # 左部矩形区域
    mask3[:y_max + pixels_far, :x_min - pixels_far] = 1  # 从左边到M区域左侧
    # 右部矩形区域
    mask4[:y_max + pixels_far, x_max + pixels_far:] = 1  # 从M区域右侧到右边
    return mask1, mask2, mask3, mask4


def save_single_image_old(image, save_path, vmin, vmax):
    plt.figure(figsize=(4, 5))
    plt.imshow(image, cmap='gray', vmin=vmin, vmax=vmax)
    # plt.title(title, fontsize=16)
    plt.axis('off')
    plt.savefig(save_path, bbox_inches='tight', dpi=400)
    plt.close()  # 关闭当前图形以释放资源


def save_single_image(mask, image, save_path, vmin, vmax):
    os.makedirs(os.path.dirname(save_path), exist_ok=True)
    # 将灰度图像归一化后转为RGB图像
    norm_image = np.clip((image - vmin) / (vmax - vmin), 0, 1)
    rgb_image = np.stack([norm_image] * 3, axis=-1)  # 变为3通道 [H,W,3]
    # 找到 mask 的边界
    boundary = find_boundaries(mask, mode='outer')  # bool数组
    # 将边界像素涂成绿色
    rgb_image[boundary] = [0, 1, 0]  # lime green, RGB=(0,255,0) in [0,1] scale
    plt.figure(figsize=(4, 5))
    plt.imshow(rgb_image, )  # 如果是灰度图像 ： cmap='gray', vmin=vmin, vmax=vmax
    plt.axis('off')
    plt.savefig(
        save_path,
        bbox_inches='tight',
        pad_inches=0,  # ❗关键：去除 padding
        dpi=400,
        facecolor='black'  # 可选：控制背景颜色，也可以用 'white'
    )
    plt.close()


def crop_subregion(img, mask, y_min, y_max, x_min, x_max, margin_y_min=0, margin_x_min=0, margin_y_max=0, margin_x_max=0, folder_name=None):
    y_min_r, y_max_r = y_min - margin_y_min, y_max + margin_y_max
    x_min_r, x_max_r = x_min - margin_x_min, x_max + margin_x_max
    # if folder_name == '2024-09-24-006_20241231_jian':
    #     print( margin_y_min, margin_x_min,margin_y_max,margin_x_max)
    #     H, W = mask.shape
    #     print(f" mask  尺寸: 高度 H = {H}（y: 0 ~ {H - 1}），宽度 W = {W}（x: 0 ~ {W - 1}）")
    #     ys, xs = np.where(mask > 0)
    #     print(f"✅ mask 中值为 1 的范围: y = {ys.min()} ~ {ys.max()}，x = {xs.min()} ~ {xs.max()}")
    #     print('裁剪范围参数',y_min_r, y_max_r,x_min_r, x_max_r)

    return img[y_min_r:y_max_r, x_min_r:x_max_r], mask[y_min_r:y_max_r, x_min_r:x_max_r]


def extract_roi_crop(ID, date_day, z, img, mask):
    if ID == '2024-09-24-006' and date_day == '20241231' and z == 73:
        plt_binary_mask(mask)
    coords = np.argwhere(mask)
    if coords.size == 0:
        raise ValueError(f"{ID, date_day, z,}掩膜中无 ROI")
    y0, x0 = coords.min(axis=0)
    y1, x1 = coords.max(axis=0) + 1
    return img[y0:y1, x0:x1], mask[y0:y1, x0:x1]


def static_pixels(valid_pixels, sub_img, mask_down):
    max_val = np.max(valid_pixels)
    min_val = np.min(valid_pixels)
    peak_to_peak = max_val - min_val
    skewness_val = skew(valid_pixels)
    kurtosis_val = kurtosis(valid_pixels)
    mean_val = np.mean(valid_pixels)
    std_val = np.std(valid_pixels)
    # 正负异常像素统计  正负异常像素 数量  比例  和正负像素能量比
    # pos_pixels = valid_pixels[valid_pixels > mean_val + 3 * std_val]
    # neg_pixels = valid_pixels[valid_pixels < mean_val - 3 * std_val]
    # 改成分位数方法
    q90 = np.percentile(valid_pixels, 90)
    q10 = np.percentile(valid_pixels, 10)
    pos_pixels = valid_pixels[valid_pixels > q90]
    neg_pixels = valid_pixels[valid_pixels < q10]
    pos_num = len(pos_pixels)
    neg_num = len(neg_pixels)
    pos_ratio = pos_num / len(valid_pixels)
    neg_ratio = neg_num / len(valid_pixels)
    # print(mean_val,std_val,neg_ratio)
    asymmetry_ratio = pos_ratio / neg_ratio if neg_ratio > 0 else np.nan
    # plt_image2(sub_img * mask_down, z, ID, date)
    # 正负像素强度面积比 = 直方图中 高于阈值像素的总强度 / 低于阈值像素的总强度 ≈ 1.0	灰度分布对称
    # > 1.0	偏高伪影（如光子硬化）
    # < 1.0	偏低伪影（如光子饥饿）
    # >> 1.0 / ≈ 0	伪影显著偏斜（高度非对称）
    asymmetry = histogram_asymmetry_ratio(sub_img, mask_down)
    return max_val, min_val, peak_to_peak, skewness_val, kurtosis_val, pos_num, neg_num, pos_ratio, neg_ratio, asymmetry_ratio, asymmetry


def mask_info(mask):
    # 🔍 打印 mask 区域信息
    ys, xs = np.where(mask > 0)
    if ys.size > 0 and xs.size > 0:
        y_min, y_max = ys.min(), ys.max() + 1
        x_min, x_max = xs.min(), xs.max() + 1
        pixel_count = np.count_nonzero(mask)
        # print(f"掩膜范围：[Y: {y_min}-{y_max}, X: {x_min}-{x_max}]")
        print(f"掩膜高度: {y_max - y_min},        宽度: {x_max - x_min}          像素数量: {pixel_count}  ")
    else:
        print("⚠掩膜为空")


def top_x_percent_mean(num, mask_combined, sub_img):
    if mask_combined.shape != sub_img.shape:
        raise ValueError("掩膜和图像的形状必须相同")
    # 获取掩膜区域内的像素值
    roi_pixels = sub_img[mask_combined == 1]
    if roi_pixels.size == 0:
        return float('nan')  # 避免掩膜区域为空时出错
    # 计算前 % 的阈值
    percentile_x = np.percentile(roi_pixels, num)  # 计算倒数最大值第 num 百分位的阈值
    # 选择像素值
    top_x_pixels = roi_pixels[roi_pixels <= percentile_x]

    if top_x_pixels.size == 0:
        return float('nan')  # 避免数组为空导致计算错误

    return top_x_pixels


def summarize_z_counts(df, metric_name):
    """
    统计每个 (Folder, ID, Date) 组内的 Z 层数，
    输出最小、最大、平均值。

    参数：
        df (pd.DataFrame): 原始数据，包含 ['Folder', 'ID', 'Date', 'Z'] 列
    返回：
        group_counts (pd.Series): 每组 Z 层数
    """
    group_counts = df.groupby(['Folder', 'ID', 'Date'])['Z'].nunique()

    # print(f"📊 {metric_name} 指标中 每个分次亚组中的计算层面数量统计：")
    print(f"   ▶ 最小层数：{group_counts.min()}")
    print(f"   ▶ 最大层数：{group_counts.max()}")
    print(f"   ▶ 平均层数：{group_counts.mean():.2f}")
    # print(f"   ▶ 总组数：{len(group_counts)}")

    return group_counts


def filter_top_metric_images(df, metric_name, image_root_dir, top_k=5):
    """
    1. 从 df 中筛选指定 metric 的 top_k Z 层（每组 Folder-ID-Date）
    2. 删除 image_root_dir 中不在这些层面的图像文件（图像命名：Folder_Z.jpg）
    3. 返回：df 中所有指标，但只保留在这些 Z 层的行（包括其他指标）

    参数：
        df (pd.DataFrame): 包含 ['Folder', 'ID', 'Date', 'Metric', 'Z', 'Value'] 等列
        metric_name (str): 要筛选的 metric 名，如 'RMSE'
        image_root_dir (str): 图像统一存放目录
        top_k (int): 每组保留前几个 Z 层
    返回：
        filtered_df (pd.DataFrame): 所有指标中在保留层面的行
    """

    # 1. 提取目标指标的数据
    df_metric = df[df['Metric'] == metric_name]

    # 2. 打印 Z 统计
    summarize_z_counts(df_metric, metric_name)

    # 3. 获取每组 Top-K 层
    print(f"\n📌 当前筛选条件：Metric = '{metric_name}', 每组保留前 {top_k} 个 Z 层。")
    top_rows = (
        df_metric
        .sort_values(['Folder', 'ID', 'Date', 'Value'], ascending=[True, True, True, False])
        .groupby(['Folder', 'ID', 'Date'])
        .head(top_k)
    )
    # ✅ 构造保留图像文件名（根据 Folder 类型分类）
    keep_images = set()
    for _, row in top_rows.iterrows():
        folder_suffix = row['Folder'].split('_')[-1]
        if folder_suffix in ['kuanbu', 'jizhu', 'yachi', 'ori']:
            # 🔁 特殊处理：用另一种方式构造
            filename = f"{folder_suffix}_{row['ID']}_{row['Date'].split('-')[0]}_{int(row['Z'])}.jpg"
        else:
            # ✅ 默认命名方式
            filename = f"{row['ID']}_{row['Date'].split('-')[0]}_{int(row['Z'])}.jpg"
        keep_images.add(filename)
    # keep_images = set(f"{row['ID']}_{row['Date'].split('-')[0]}_{int(row['Z'])}.jpg" for _, row in top_rows.iterrows())

    # print("🖼️ 保留图像文件名示例：", list(keep_images)[:3])

    # 5. 清理图像
    fname_list = []
    for fname in os.listdir(image_root_dir):
        if fname.endswith('.jpg') and fname not in keep_images:
            os.remove(os.path.join(image_root_dir, fname))
            fname_list.append(fname)
    print(f"🗑️ 已删除：{fname_list}")

    # 6. 构造保留的 key 用于筛选所有指标
    retained_keys = set(zip(top_rows['Folder'], top_rows['ID'], top_rows['Date'], top_rows['Z']))

    # 7. 在全 df 中保留这些 key 的行（包含所有指标）
    filtered_df = df[
        df.apply(lambda row: (row['Folder'], row['ID'], row['Date'], row['Z']) in retained_keys, axis=1)
    ]

    # print(f"\n✅ 筛选完成：已保留所有指标中与 '{metric_name}' 前 {top_k} 层匹配的 Z 层面数据。")
    return filtered_df


def save_dataframe_with_filter(df, excel_path, sheet_name="Sheet1"):
    """
    保存 DataFrame 到 Excel，并添加筛选功能（AutoFilter）
    """
    # 1. 首先用 pandas 保存 Excel 文件
    df.to_excel(excel_path, index=False, sheet_name=sheet_name)

    # 2. 用 openpyxl 加载刚保存的文件并打开表格
    wb = load_workbook(excel_path)
    ws = wb[sheet_name]

    # 3. 获取表格范围，比如 A1:G100
    max_row = ws.max_row
    max_col = ws.max_column
    col_letter = chr(64 + max_col) if max_col <= 26 else None  # 简单支持 A-Z
    if not col_letter:
        from openpyxl.utils import get_column_letter
        col_letter = get_column_letter(max_col)
    table_range = f"A1:{col_letter}{max_row}"

    # 4. 添加表格对象（自动带筛选）
    table = Table(displayName="FilteredTable", ref=table_range)

    # 5. 添加样式（可选）
    style = TableStyleInfo(
        name="TableStyleMedium9", showFirstColumn=False,
        showLastColumn=False, showRowStripes=True, showColumnStripes=False
    )
    table.tableStyleInfo = style

    ws.add_table(table)
    wb.save(excel_path)
    # print(f"✅ Excel 文件保存并启用筛选功能：{excel_path}")


def get_slice_mask(z, volume, slice_mask):
    slice_img = volume[z, :, :].copy()
    # print(slice_img)
    # 默认先裁剪一个最小外切矩形用于计算
    # 找到结构的最小外接矩形区域  max位置加1 因为裁剪时是右边是不包含该值的关系
    ys, xs = np.where(slice_mask > 0)
    y_min, y_max = ys.min(), ys.max() + 1
    x_min, x_max = xs.min(), xs.max() + 1
    return slice_img, y_min, y_max, x_min, x_max


def cal_radius(sub_img, spacing, th=3000):
    # 1. 提取图像中 >3000 HU 的区域
    binary_mask = (sub_img > th).astype(np.uint8)
    # 2. 计算区域内的像素总数
    num_pixels = np.sum(binary_mask)
    # 3. 计算面积（mm²）
    pixel_area_mm2 = spacing[0] * spacing[1]  # 注意 spacing 是 [row_spacing, col_spacing]
    area_mm2 = num_pixels * pixel_area_mm2
    # 4. 计算等效圆的半径 r = sqrt(A / π)
    if area_mm2 > 0:
        radius_mm = np.sqrt(area_mm2 / np.pi)
    else:
        radius_mm = 0
    return radius_mm


def visualize_key_rows_cols(slice_mask, image, key_rows, key_cols, title='关键行列索引', save_path=None):
    """
    在灰度图像上叠加半透明红色线，标记关键行列索引。
    image: 2D 图像 (灰度)
    key_rows, key_cols: 列表[int]
    """
    # 1. 复制图像并转为 RGB 显示格式
    img_rgb = np.stack([image] * 3, axis=-1)  # shape: (H, W, 3)

    # 2. mask 区域设为
    if slice_mask is not None:
        img_rgb[slice_mask > 0] = [150, 150, 150]

    # 3. 绘制关键行列
    if key_rows:
        for y in key_rows:
            if 0 <= y < img_rgb.shape[0]:
                img_rgb[y, :, :] = [255, 100, 100]
    if key_cols:
        for x in key_cols:
            if 0 <= x < img_rgb.shape[1]:
                img_rgb[:, x, :] = [255, 100, 100]

    # 4. 绘图展示
    plt.figure(figsize=(6, 6), dpi=300)
    plt.imshow(img_rgb.astype(np.uint8), vmin=0, vmax=255)
    plt.title(title)
    plt.axis('off')
    if save_path:
        plt.savefig(save_path, dpi=300, bbox_inches='tight')
        plt.close()
        print(f'✔️ 图像已保存：{save_path}')
    else:
        plt.show()


def cal_diameter(ID, date, z, roi, sub_img, slice_mask, spacing, th, labeled, labels_to_use):
    """
        识别掩膜中的三个施源器截面（连通域），仅取最左与次左两管，
        在中心 3×3 像素块内统计 HU>th 的像素并转 mm。
        返回 dict: row_mm, col_mm, row_counts, col_counts
        """
    dy, dx = spacing[0], spacing[1]  # dy: 行 mm/px, dx: 列 mm/px
    # 获取掩膜mask的xy最大范围
    ys, xs = np.where(slice_mask > 0)  # 找到非零像素的行、列索引
    y_min, y_max = ys.min(), ys.max()
    x_min, x_max = xs.min(), xs.max()
    # # 1) 连通域标记
    # labeled, n_cc = ndi.label(slice_mask)
    # if n_cc != 3:
    #     print(f"检测到{ID,date,z,roi,metric} 存在{n_cc} 个连通域 (最好为3)，请检查掩膜")
    # # 2) 计算每个连通域的质心 X 坐标并排序
    # cc_props = ndi.center_of_mass(slice_mask, labeled, range(1, n_cc + 1))
    # cc_sorted = sorted(enumerate(cc_props, start=1), key=lambda x: x[1][1])  # 根据X坐标排序
    # 取第1、3个施源器连通域 label
    # # 3) 获取最多前两个（最左和最右），跳过中间段   目前已经在外部只计算连通区域为3的层面了
    # if n_cc == 3:
    #     labels_to_use = [cc_sorted[0][0], cc_sorted[2][0]]
    # elif n_cc == 2:
    #     # labels_to_use = [cc_sorted[0][0],cc_sorted[1][0]]
    #     raise ValueError(f"❌ 检测到2金属区域，不确定哪个是左右两侧的施源器 请检查 {ID}, {date}, z={z}, ROI={roi}")
    # elif n_cc == 1:
    #     labels_to_use = [cc_sorted[0][0]]
    # else:
    #     raise ValueError(f"❌ 无法检测到任何有效金属区域，请检查 {ID}, {date}, z={z}, ROI={roi}")
    row_counts, col_counts = [], []
    key_rows, key_cols = [], []
    for lab, i in zip(labels_to_use, [1, 2, 3]):
        ys, xs = np.where(labeled == lab)
        # 获取当前连通区域的范围
        ys_min, ys_max = ys.min(), ys.max()
        xs_min, xs_max = xs.min(), xs.max()
        y_c = int(np.round(np.mean(ys)))
        x_c = int(np.round(np.mean(xs)))
        # 各取中心的5行和5列
        # rows3 = [y for y in (y_c - 2,y_c - 1, y_c, y_c + 1,y_c + 2) if y_min <= y <= y_max]
        # cols3 = [x for x in (x_c - 2,x_c - 1, x_c, x_c + 1,x_c + 2) if x_min <= x <= x_max]
        # rows3 = [y_c]
        # cols3 = [x_c]
        rows3 = [y for y in (y_c - 1, y_c, y_c + 1) if y_min <= y <= y_max]
        cols3 = [x for x in (x_c - 1, x_c, x_c + 1) if x_min <= x <= x_max]
        key_rows.extend(rows3)
        key_cols.extend(cols3)
        # 行向统计（对选行求均值）
        if len(labels_to_use) == 2 and i == 1:  # 代表遍历的是第一个 也就是最左边的连通区域 这里要注意计算的列范围为最左边到连通区域右边一定范围。 否则可能会不小心计算到下一个连通区域内的高像素值
            pix_row = [np.sum(sub_img[y, :xs_max + 2] > th) for y in rows3]
        elif len(labels_to_use) == 2 and i == 2:  # 代表是最右边的连通区域
            pix_row = [np.sum(sub_img[y, xs_min - 2:] > th) for y in rows3]
        elif len(labels_to_use) == 1 and i == 1:  # 代表估计是单管层面 不存在三管 那么就计算每一行的左右两侧所有
            pix_row = [np.sum(sub_img[y, :] > th) for y in rows3]
        else:
            raise ValueError(f"❌ 检测到{len(labels_to_use)} 连通域，当前遍历第{i}.  请检查 {ID}, {date}, z={z}, ROI={roi}")
        # if roi == 'spc1':
        #     print(f'{ID,date,z,roi,metric,th}  第{i}个连通域  共{len(rows3)}行的阈值以上像素个数', pix_row)
        row_counts.append(np.mean(pix_row))
        # 列向统计（对选列求均值）
        pix_col = [np.sum(sub_img[:, x] > th) for x in cols3]
        # if roi == 'spc1':
        #     print(f'{ID,date,z,roi,metric,th}  第{i}个连通域  共{len(cols3)}列的阈值以上像素个数', pix_col)
        col_counts.append(np.mean(pix_col))
    # 平均像素个数
    row_counts = np.array(row_counts)
    col_counts = np.array(col_counts)
    # 行列平均物理长度
    row_mm = np.mean(row_counts) * dx
    col_mm = np.mean(col_counts) * dy
    # if roi == 'spc1':
    #     # print('行索引包括', key_rows, '列索引包括', key_cols, 'mask的xy范围为', y_min, y_max, x_min, x_max)
    #     print(f"{ID,date,z,roi,metric,th} 行平均像素个数为 {row_counts}，列平均像素个数为 {col_counts}")
    #     print(f"{ID,date,z,roi,metric,th} 行平均物理长度为 {row_mm:.2f} mm, 列平均物理长度为 {col_mm:.2f} mm")
    # 可视化检查
    # visualize_key_rows_cols(slice_mask,image=sub_img,key_rows=key_rows,key_cols= key_cols,title='验证行列索引')
    # input('下一个')
    return row_mm, col_mm


def normalize_metric_values(df, group_keys, method='min-max', value_col='Value', new_col='group_norm_Value', round_ndigits=4):
    """
    对 DataFrame 中的 'Value' 列按指定分组进行归一化，并生成新列。
    参数：
        df: 输入 DataFrame，需包含 'Value' 列。
        group_keys: 列表，指定按哪些列分组，例如 ['Folder', 'ID', 'Date', 'ROI', 'Metric']
        method: 归一化方法，可选：'min-max', 'z-score', 'max-abs'
        value_col: 原始值列名（默认是 'Value'）
        new_col: 新列名（默认是 'NormValue'）
        round_ndigits: 保留的小数位数（默认 4）

    返回：
        新增归一化列的 DataFrame
    """

    def normalize_group(group):
        values = group[value_col].astype(float).values
        if method == 'min-max':
            vmin, vmax = values.min(), values.max()
            if vmax != vmin:
                group[new_col] = (values - vmin) / (vmax - vmin)
            else:
                group[new_col] = 0.0
        elif method == 'z-score':
            mean, std = values.mean(), values.std()
            if std != 0:
                group[new_col] = (values - mean) / std
            else:
                group[new_col] = 0.0
        elif method == 'max-abs':
            vmax = np.max(np.abs(values))
            if vmax != 0:
                group[new_col] = values / vmax
            else:
                group[new_col] = 0.0
        else:
            raise ValueError(f"Unsupported normalization method: {method}")
        group[new_col] = group[new_col].round(round_ndigits)
        return group

    df = df.groupby(group_keys, group_keys=False).apply(normalize_group)
    return df


def compute_multi_reference_metrics_grouped(
        ct_results,
        roi_list=('target_roi',),
        ref_keywords=('ori',),
        metrics=('ncc', 'psnr', 'ssim', 'rmse')
):
    """对同一 ID-Date 组的 CT 图像，计算多个参考图像的质量指标。
       结果存储在：ct_results[target][roi][metric][z][ref] = value
    """
    # ──────────────────────────────────────────────
    # 1) 先把 Folder 按 (ID, date) 分组
    grouped_folders = defaultdict(list)
    for folder, info in ct_results.items():
        grouped_folders[(info.get('ID'), info.get('date'))].append(folder)
    # ──────────────────────────────────────────────
    # 2) 遍历每组，生成 (ref, target) 配对
    for (_id, _date), folders in grouped_folders.items():
        if ref_keywords:  # 给定参考图像的标签 比如所有包含ori字符串的文件夹 作为参考文件夹
            ref_folders = [f for f in folders if any(k in f for k in ref_keywords)]
            non_ref_folders = [f for f in folders if f not in ref_folders]
            pairs = [(ref, tgt) for ref in ref_folders for tgt in non_ref_folders]
        else:  # 全两两配对
            # 不考虑顺序的：每幅图都当参考但分别位于ref-image列和folder列” → 相对于两两组合36种
            # pairs = [(f1, f2) for i, f1 in enumerate(folders) for f2 in folders[i + 1:]]
            # 考虑顺序的： 每幅图都当参考且位于ref-image列” → 就自然有 9 个组，每组 8 行 的结果 72个  不包含自己比自己
            # pairs = [(ref, tgt) for ref in folders for tgt in folders if ref != tgt]
            # 不仅包含顺序 还包含自己比较自己的指标 例如ncc =1
            pairs = [(ref, tgt) for ref in folders for tgt in folders]
        # ──────────────────────────────────────────
        for ref_folder, target_folder in pairs:
            # print(f"当前遍历 {_id, _date} {ref_folder, target_folder}")
            for roi in roi_list:
                # print( f"当前遍历 {_id, _date} {ref_folder, target_folder} {roi}")
                if roi not in ct_results[ref_folder] or roi not in ct_results[target_folder]:
                    continue
                # 取两者共同的 z
                z_vals = sorted(
                    set(ct_results[ref_folder][roi].get('zcoord', {})).intersection(
                        ct_results[target_folder][roi].get('zcoord', {})))
                # 遍历指标 & z
                for metric in metrics:
                    tgt_roi_metric = ct_results[target_folder][roi].setdefault(metric, {})
                    for z in z_vals:
                        img_ref = ct_results[ref_folder][roi].get('sub_images', {}).get(z)
                        img_tgt = ct_results[target_folder][roi].get('sub_images', {}).get(z)
                        if img_ref is None or img_tgt is None or img_ref.shape != img_tgt.shape:
                            continue
                        # 计算 value
                        if metric == 'ncc':
                            value = np.corrcoef(img_ref.ravel(), img_tgt.ravel())[0, 1]
                        elif metric == 'psnr':
                            value = peak_signal_noise_ratio(img_ref, img_tgt,
                                                            data_range=img_ref.max() - img_ref.min())
                        elif metric == 'ssim':
                            img_ref2 = resize(img_ref, (64, 64), preserve_range=True,
                                              anti_aliasing=True, mode='reflect')
                            img_tgt2 = resize(img_tgt, (64, 64), preserve_range=True,
                                              anti_aliasing=True, mode='reflect')
                            value = structural_similarity(img_ref2, img_tgt2,
                                                          data_range=img_ref.max() - img_ref.min())
                        elif metric == 'rmse':
                            value = np.sqrt(mean_squared_error(img_ref, img_tgt))
                        else:
                            continue
                        # 写入：metric[z] 是 dict， 仍以 ref_folder 为键，避免覆盖  这里相当于无参考指标结果中的noref键
                        # 'add_tag': 'no'  # 可添加特殊的标签 同时与无参考指标格式的赋值函数中 一致
                        write_metric(ct_results, target_folder, roi, metric, z, value, ref=ref_folder)
    return ct_results


def generate_parameter_groups_for(
        y_range=(2, 9), y_step=2,
        x_range=(4, 21), x_step=2,
        inner_range=(1, 8)
):
    """
    根据给定范围与步长生成合法的参数组合（外边界 ≥ 内边界），格式：
    [margin_y_min, margin_x_min, margin_y_max, margin_x_max, inner, '参数组编号']
    参数：
    - y_range: tuple，y方向边界的最小值与最大值（不含最大值），如 (2, 9)
    - y_step: int，y方向步长
    - x_range: tuple，x方向边界的最小值与最大值（不含最大值），如 (4, 21)
    - x_step: int，x方向步长
    - inner_range: tuple，inner值范围，如 (1, 8)
    返回：
    - List[List]，每项是一个合法参数组合
    """
    parameters_list = []
    count = 1
    for margin_y in range(*y_range, y_step):
        for margin_x in range(*x_range, x_step):
            for inner in range(*inner_range):
                if margin_y > inner and margin_x > inner:
                    param_group = [
                        margin_y, margin_x, margin_y, margin_x,
                        inner, f'param_{count}'
                    ]
                    parameters_list.append(param_group)
                    count += 1
    return parameters_list


def far_from_mask(mask_3d_spc1, slice_mask, z, pixel_far):
    slice_mask_spc1 = mask_3d_spc1[z]  # 因为先遍历spc1  所以一定是有这个结构的
    rows_spc1 = np.where(slice_mask_spc1.any(axis=1))[0]
    y0 = rows_spc1.min() - pixel_far
    y1 = rows_spc1.max() + pixel_far
    slice_mask[y0:y1 + 1, :] = 0
    return slice_mask


def correct_and_crop_region_for_cal_spc1_MAS(folder_name, mask_3d, ID, date_day, z, volume, slice_img, slice_mask, y_min_roi, y_max_roi, x_min_roi, x_max_roi, margin_y_min, margin_x_min, margin_y_max, margin_x_max):
    # print(f"正在处理 {folder_name} {ID} {date_day} {z}")
    # print(f"margin_y_min: {margin_y_min}, margin_x_min: {margin_x_min}, margin_y_max: {margin_y_max}, margin_x_max: {margin_x_max}")
    if ID == '2024-08-09-004':  # and (date_day == '20250107' or date_day == '20250110')
        sub_img, slice_mask2 = crop_subregion(slice_img, slice_mask, y_min_roi, y_max_roi, x_min_roi, x_max_roi,
                                              margin_y_min=min(8, margin_y_min), margin_x_min=min(17, margin_x_min),
                                              margin_y_max=min(8, margin_y_max), margin_x_max=min(14, margin_x_max))
    elif ID == '2025-02-21-022' and date_day == '20250228' and z == 46:
        slice_mask = mask_3d[45]
        slice_img, y_min, y_max, x_min, x_max = get_slice_mask(45, volume, slice_mask)
        sub_img, slice_mask2 = crop_subregion(slice_img, slice_mask, y_min_roi, y_max_roi, x_min_roi, x_max_roi,
                                              margin_y_min=min(8, margin_y_min), margin_x_min=min(20, margin_x_min),
                                              margin_y_max=min(8, margin_y_max), margin_x_max=min(20, margin_x_max))
    elif ID == '2025-02-21-022' and date_day == '20250225':
        sub_img, slice_mask2 = crop_subregion(slice_img, slice_mask, y_min_roi, y_max_roi, x_min_roi, x_max_roi,
                                              margin_y_min=min(8, margin_y_min), margin_x_min=min(20, margin_x_min),
                                              margin_y_max=min(10, margin_y_max), margin_x_max=min(20, margin_x_max))
    elif ID == '2024-11-05-012' and date_day == '20250307' and z == 60:
        sub_img, slice_mask2 = crop_subregion(slice_img, slice_mask, y_min_roi, y_max_roi, x_min_roi, x_max_roi,
                                              margin_y_min=min(8, margin_y_min), margin_x_min=min(20, margin_x_min),
                                              margin_y_max=min(8, margin_y_max), margin_x_max=min(20, margin_x_max))
    elif ID == '2024-11-05-012' and date_day == '20250307' and z != 60:
        sub_img, slice_mask2 = crop_subregion(slice_img, slice_mask, y_min_roi, y_max_roi, x_min_roi, x_max_roi,
                                              margin_y_min=min(8, margin_y_min), margin_x_min=min(20, margin_x_min),
                                              margin_y_max=min(8, margin_y_max), margin_x_max=min(20, margin_x_max))
    elif ID == '2024-11-05-012' and date_day == '20250218' and z == 55 or z == 56:
        sub_img, slice_mask2 = crop_subregion(slice_img, slice_mask, y_min_roi, y_max_roi, x_min_roi, x_max_roi,
                                              margin_y_min=min(8, margin_y_min), margin_x_min=min(20, margin_x_min),
                                              margin_y_max=min(8, margin_y_max), margin_x_max=min(20, margin_x_max))
    elif ID == '2024-11-05-012' and date_day == '20250218' and z == 57:
        sub_img, slice_mask2 = crop_subregion(slice_img, slice_mask, y_min_roi, y_max_roi, x_min_roi, x_max_roi,
                                              margin_y_min=min(8, margin_y_min), margin_x_min=min(20, margin_x_min),
                                              margin_y_max=min(8, margin_y_max), margin_x_max=min(20, margin_x_max))
    elif ID == '2024-11-05-012' and date_day == '20250221' and z == 54:
        sub_img, slice_mask2 = crop_subregion(slice_img, slice_mask, y_min_roi, y_max_roi, x_min_roi, x_max_roi,
                                              margin_y_min=min(8, margin_y_min), margin_x_min=min(20, margin_x_min),
                                              margin_y_max=min(8, margin_y_max), margin_x_max=min(20, margin_x_max))
    elif ID == '2024-11-05-012' and date_day == '20250221' and z == 55:
        sub_img, slice_mask2 = crop_subregion(slice_img, slice_mask, y_min_roi, y_max_roi, x_min_roi, x_max_roi,
                                              margin_y_min=min(8, margin_y_min), margin_x_min=min(20, margin_x_min),
                                              margin_y_max=min(8, margin_y_max), margin_x_max=min(20, margin_x_max))
    elif ID == '2024-07-15-008' and date_day == '20241220' and z == 55:
        sub_img, slice_mask2 = crop_subregion(slice_img, slice_mask, y_min_roi, y_max_roi, x_min_roi, x_max_roi,
                                              margin_y_min=min(8, margin_y_min), margin_x_min=min(20, margin_x_min),
                                              margin_y_max=min(8, margin_y_max), margin_x_max=min(20, margin_x_max))
    elif ID == '2024-07-15-008' and date_day == '20241220' and z == 54:
        sub_img, slice_mask2 = crop_subregion(slice_img, slice_mask, y_min_roi, y_max_roi, x_min_roi, x_max_roi,
                                              margin_y_min=min(8, margin_y_min), margin_x_min=min(20, margin_x_min),
                                              margin_y_max=min(8, margin_y_max), margin_x_max=min(20, margin_x_max))
    elif ID == '2024-11-06-002' and date_day == '20250304':
        sub_img, slice_mask2 = crop_subregion(slice_img, slice_mask, y_min_roi, y_max_roi, x_min_roi, x_max_roi,
                                              margin_y_min=min(8, margin_y_min), margin_x_min=min(20, margin_x_min),
                                              margin_y_max=min(2, margin_y_max), margin_x_max=min(20, margin_x_max))
    elif ID == '2024-09-24-006' and date_day == '20241231' and (z == 73 or z == 74):
        # print(1)
        slice_mask = mask_3d[62]
        slice_img, y_min, y_max, x_min, x_max = get_slice_mask(62, volume, slice_mask)
        print(y_min, y_max, x_min, x_max)
        plt_binary_mask(slice_mask)
        print(margin_y_min, margin_y_max, margin_x_max, margin_x_min)
        y_min_r, y_max_r = y_min - margin_y_min, y_max + margin_y_max
        x_min_r, x_max_r = x_min - margin_x_min, x_max + margin_x_max
        print(y_min_r, y_max_r, x_min_r, x_max_r)
        sub_img, slice_mask2 = crop_subregion(slice_img, slice_mask, y_min_roi, y_max_roi, x_min_roi, x_max_roi,
                                              margin_y_min=min(8, margin_y_min), margin_x_min=min(20, margin_x_min),
                                              margin_y_max=min(8, margin_y_max), margin_x_max=min(20, margin_x_max))
        # plt_image_with_mask_and_save(sub_img, slice_mask2, ID=ID, date=date_day, z=z)
        plt_binary_mask(slice_mask2)
        plt_image2(sub_img, ID=ID, date=date_day, z=z)

    elif ID == '2024-09-24-006' and date_day == '20241226' and (z == 72 or z == 73):
        slice_mask = mask_3d[72]
        slice_img, y_min, y_max, x_min, x_max = get_slice_mask(72, volume, slice_mask)
        sub_img, slice_mask2 = crop_subregion(slice_img, slice_mask, y_min_roi, y_max_roi, x_min_roi, x_max_roi,
                                              margin_y_min=min(8, margin_y_min), margin_x_min=min(20, margin_x_min),
                                              margin_y_max=min(8, margin_y_max), margin_x_max=min(20, margin_x_max))
    elif ID == '2025-01-08-011' and date_day == '20250121' and z == 57:
        slice_mask = mask_3d[54]
        slice_img, y_min, y_max, x_min, x_max = get_slice_mask(54, volume, slice_mask)
        sub_img, slice_mask2 = crop_subregion(slice_img, slice_mask, y_min_roi, y_max_roi, x_min_roi, x_max_roi,
                                              margin_y_min=min(8, margin_y_min), margin_x_min=min(20, margin_x_min),
                                              margin_y_max=min(8, margin_y_max), margin_x_max=min(20, margin_x_max))
    elif ID == '2024-12-02-018' and date_day == '20250225' and z == 48:
        sub_img, slice_mask2 = crop_subregion(slice_img, slice_mask, y_min_roi, y_max_roi, x_min_roi, x_max_roi,
                                              margin_y_min=min(8, margin_y_min), margin_x_min=min(20, margin_x_min),
                                              margin_y_max=min(8, margin_y_max), margin_x_max=min(12, margin_x_max))
    elif ID == '2024-12-02-018' and date_day == '20250304' and z == 45:
        sub_img, slice_mask2 = crop_subregion(slice_img, slice_mask, y_min_roi, y_max_roi, x_min_roi, x_max_roi,
                                              margin_y_min=min(8, margin_y_min), margin_x_min=min(20, margin_x_min),
                                              margin_y_max=min(8, margin_y_max), margin_x_max=min(12, margin_x_max))
    elif ID == '2024-12-02-018' and date_day == '20250228' and z in [36, 37, 38]:
        slice_mask = mask_3d[40]
        slice_img, y_min, y_max, x_min, x_max = get_slice_mask(40, volume, slice_mask)
        sub_img, slice_mask2 = crop_subregion(slice_img, slice_mask, y_min_roi, y_max_roi, x_min_roi, x_max_roi,
                                              margin_y_min=margin_y_min, margin_x_min=margin_x_min,
                                              margin_y_max=margin_y_max, margin_x_max=margin_x_max)
    elif ID == '2024-11-08-013' and date_day == '20250211' and z in [49, 50]:
        sub_img, slice_mask2 = crop_subregion(slice_img, slice_mask, y_min_roi, y_max_roi, x_min_roi, x_max_roi,
                                              margin_y_min=min(margin_y_min, 5), margin_x_min=margin_x_min,
                                              margin_y_max=margin_y_max,
                                              margin_x_max=margin_x_max)  # 反正就是要在75754参数下再 y min 减两格注意方向
    else:
        sub_img, slice_mask2 = crop_subregion(slice_img, slice_mask, y_min_roi, y_max_roi, x_min_roi, x_max_roi,
                                              margin_y_min=margin_y_min, margin_x_min=margin_x_min,
                                              margin_y_max=margin_y_max, margin_x_max=margin_x_max, folder_name=folder_name)
        # plt_image_with_mask_and_save(sub_img, slice_mask2, ID=ID, date=date_day, z=z)
        # exit()
    # if folder_name == '2024-09-24-006_20241231_jian':
    #     H, W = slice_mask2.shape
    #     print(f" slice_mask2  尺寸: 高度 H = {H}（y: 0 ~ {H - 1}），宽度 W = {W}（x: 0 ~ {W - 1}）")
    return sub_img, slice_mask2


def write_metric(ct_results, folder, roi, metric, z, value, ref='noref'):  # tag='no',
    ct_results \
        .setdefault(folder, {}) \
        .setdefault(roi, {}) \
        .setdefault(metric, {}) \
        .setdefault(z, {})[ref] = {
        'value': value,
        # 'add_tag': tag  # 可以视需求 添加额外的tag 标签  并在函数参数中给这个参数
    }


def correct_pixel_for_MAS_cal_in_spc1(slice_mask, slice_img):
    th_crop = -40
    th_crop2 = -100  # 这里是指ori中光子饥饿的低HU不要错误提高了  需要测试  大概就是-100 左右
    # 只处理 mask 即spc1施源器 之外的区域
    outside_mask = (slice_mask == 0)
    # 提取 mask 外的像素视图（避免重复索引）
    outside_vals = slice_img[outside_mask]
    # 分段处理：按顺序从最小到最大处理替换区间
    outside_vals[(outside_vals >= th_crop2) & (
            outside_vals < th_crop)] = -10  # 这里特殊处理 意图是让软组织变成和周围高软组织差不多的 这样就只有伪影的影响了 不要设为0 因为我计算直径最小阈值为0 会受影响
    # 赋回图像
    slice_img[outside_mask] = outside_vals
    return slice_img


def get_sub_img_by_mask(img, mask):
    ys, xs = np.where(mask > 0)
    y_min, y_max = ys.min(), ys.max() + 1
    x_min, x_max = xs.min(), xs.max() + 1
    sub_img = img[y_min:y_max, x_min:x_max]
    return sub_img


def save_subroi_to_ct_results(ct_results, folder_name, base_roi, suffix, tag, z, sub_img, mask=None):
    """
    将子 ROI 的图像与掩膜保存到 ct_results 中，以新的 ROI 命名（base_roi_suffix_tag）
    参数说明：
    - ct_results: 总字典
    - folder_name: 当前病例文件夹名
    - base_roi: 原始 ROI 名（如 spc1）
    - suffix: 参数组合标识符（如 '040408043'）
    - tag: 方向标记（如 'up', 'down'）
    - z: 层号（切片索引）
    - sub_img: 子图像（np.ndarray）
    - mask: 对应 mask（np.ndarray），可选
    """
    new_roi = f"{base_roi}_{suffix}_{tag}"
    roi_dict = ct_results.setdefault(folder_name, {}).setdefault(new_roi, {})
    # 🔒 防御性编程
    if not isinstance(roi_dict, dict):
        raise TypeError(
            f"[致命错误] roi_dict 不是 dict，而是 {type(roi_dict)}，说明 ct_results[folder_name][base_roi] 被污染了。")
    roi_dict.setdefault('sub_images', {})[z] = sub_img.astype(np.float32)


def generate_ring_mask(mask_tags, mask, dilation_iter=3, erosion_iter=3):
    mask_dict = {}
    # 膨胀：生成“更大”的区域
    dilated = cv2.dilate(mask, None, iterations=dilation_iter)
    # 腐蚀：生成“更小”的区域
    eroded = cv2.erode(mask, None, iterations=erosion_iter)
    # 环形 = 膨胀结果 - 腐蚀结果
    ring = cv2.subtract(dilated, eroded)
    # 找到 mask 有效区域的最小和最大行号
    rows = np.any(mask, axis=1)
    cols = np.any(mask, axis=0)
    y_min, y_max = np.where(rows)[0][[0, -1]]
    x_min, x_max = np.where(cols)[0][[0, -1]]
    y_mid = (y_min + y_max) // 2
    x_mid = (x_min + x_max) // 2
    for tag in mask_tags:
        # if tag == 'rectum_ring_l':
        #     ring_l = ring.copy()
        #     ring_l[:, x_mid + 1:] = 0
        #     mask_dict[tag] = ring_l
        # elif tag == 'rectum_ring_r':
        #     ring_r = ring.copy()
        #     ring_r[:, :x_mid] = 0
        #     mask_dict[tag] = ring_r
        # elif tag == 'rectum_ring_l_half': # 只保留掩膜区域的上半部分的 ring
        #     ring_l2 = ring.copy()
        #     ring_l2[y_mid + 1:, x_mid + 1:] = 0
        #     mask_dict[tag] = ring_l2
        # elif tag == 'rectum_ring_r_half':
        #     ring_r2 = ring.copy()
        #     ring_r2[y_mid + 1:, :x_mid] = 0
        #     mask_dict[tag] = ring_r2
        if tag == 'rectum_ring_LCR':
            mask_dict[tag] = ring
        # elif tag == 'ring_include_boundry':  # 这个是生成了带有直肠边界的整体环 不分左右 以便后续计算特殊指标 即环直肠外均值减去环直肠内均值等等
        #     ring_inner = cv2.subtract(mask, eroded)  # 它包含原始mask边界
        #     ring_outer = cv2.subtract(dilated, mask)
        #     ring_inner[ring_outer > 0] = 2
        #     ring_inner[y_mid + 1:] = 0  # 只保留掩膜区域的上半部分的 和原始mask的顶部以上就不要外扩了
        #     ring_inner[:y_min] = 0
        #     mask_dict[tag] = ring_inner
        #     # 看情况要不要再分左右
        #     # ring_inner_l = ring_inner.copy()
        #     # ring_inner_l[ x_mid + 1:] = 0
        #     # mask_dict[f'{tag}_l'] = ring_inner
        #     # mask_dict[f'{tag}_r'] = ring_inner
    return mask_dict


def generate_dilated_upper_half_mask(mask, dilation_iter=3, propotion=2 / 3):
    # 膨胀：生成“更大”的区域
    dilated = cv2.dilate(mask, None, iterations=dilation_iter)
    # 保留上propotion部分
    rows = np.any(mask, axis=1)
    y_min, y_max = np.where(rows)[0][[0, -1]]
    y_propotion = int(y_min + (y_max - y_min) * propotion)
    # 限制膨胀区域不超过原始mask的上边界（y_min）
    dilated[:y_min, :] = 0
    # 只保留掩膜区域的上半部分的 ring
    dilated[y_propotion + 1:, :] = 0
    return dilated.astype(np.uint8)


def generate_two_param_sets(para1, para2):
    param_sets = []
    for d in para1:
        for e in para2:
            param_sets.append([d, e, None, None, None])
    return param_sets


def process_ID_of_rectum_mask(date_day, ID, z, add_para, slice_mask2):
    # 特殊ID 靠近施源器部分 给赋值0
    if date_day == '20250225' and ID == '2025-02-21-022':
        slice_mask2 = far_from_mask(add_para, slice_mask2, z, 9)
        if not np.any(slice_mask2):
            print(f'{ID, date_day, z} 没有mask')
    elif date_day == '20250411' and ID == '2025-01-02-010':
        slice_mask2 = far_from_mask(add_para, slice_mask2, z, 8)
    elif date_day == '20250124' and ID == '2024-10-28-011':
        slice_mask2 = far_from_mask(add_para, slice_mask2, z, 8)
    return slice_mask2


# =======================================================================
# 1) 通用派生-ROI 处理函数（放在 for-z 循环**之前**定义一次即可）
# =======================================================================
def gen_subroi_masks(mask_3d, ID, date_day, z, volume, ct_results, folder_name, base_roi, para_name, parameters,
                     slice_img, slice_mask,
                     y_min_roi, y_max_roi, x_min_roi,
                     x_max_roi,
                     mask_tags, add_para=''):  # add_para参数被给定为特殊额外的变量 对于直肠extend就是 spc1的mask
    """
    统一派生 ROI，
    并通过 save_subroi_to_ct_results() 写回 ct_results，返回 mask_dict
    """
    mask_dict = OrderedDict()
    margin_param_list = parameters[para_name]
    # if ID == '2024-09-24-006' and date_day == '20241231' and z == 73:
    #     plt_binary_mask(slice_mask)

    # ---------- 1. 可选灰度校正 ----------
    # print('本次未对spc1_for_MAS做灰度校正')
    if 'spc1_for_MAS' == para_name:  # 对于spc1计算MAS指标的图像：将特定范围阈值以下的HU都设为阈值 因为施源器周围很多脂肪组织和伪影高像素无关 因此保存的时候也保存处理后的图像看看效果
        slice_img = correct_pixel_for_MAS_cal_in_spc1(slice_mask, slice_img)
    for one_param in margin_param_list:
        # 解包参数
        my, mx, My, Mx, pixels_far = one_param[:5]
        suffix_prefix = f"{my}{round(mx, 2)}{My}{Mx}{pixels_far}"
        if para_name in ['spc1_for_MAS', 'spc1_for_LCR', 'HRCTV_for_LCR', 'rectum_for_LCR']:
            if 'spc1_for_MAS' == para_name:
                sub_img, slice_mask2 = correct_and_crop_region_for_cal_spc1_MAS(folder_name,
                                                                                mask_3d, ID, date_day, z, volume, slice_img, slice_mask, y_min_roi, y_max_roi, x_min_roi, x_max_roi, my, mx, My, Mx)
            else:
                # 通过my, mx, My, Mx,参数 生成roi扩大后对应的sub_img, slice_mask2   然后在通过 pixels_far 裁剪出不同mask
                sub_img, slice_mask2 = crop_subregion(slice_img, slice_mask, y_min_roi, y_max_roi, x_min_roi, x_max_roi, margin_y_min=my, margin_x_min=mx, margin_y_max=My, margin_x_max=Mx)

            # if ID == '2024-09-24-006' and date_day == '20241231' and z == 73:
            #     # plt_binary_mask(slice_mask2)
            #     plt_image_with_mask_and_save(sub_img, slice_mask2, ID=ID, date=date_day, z=z)
            masks_dict = create_remaining_masks(folder_name, slice_mask2, pixels_far, mask_tags, para_name=para_name)
            subroi_dict = {}
            for tag in mask_tags:
                if tag not in masks_dict:
                    continue
                m = masks_dict[tag]
                sub_img_d, mask_d = extract_roi_crop(ID, date_day, z, sub_img, m)  # 该函数将大mask 和subimg 缩小为最小外接矩形的形式
                save_subroi_to_ct_results(ct_results, folder_name, base_roi,
                                          suffix_prefix, tag, z, sub_img_d, mask_d)
                subroi_dict[tag] = {'img': sub_img_d.astype(np.float32),
                                    'mask': mask_d.astype(np.uint8)}
            mask_dict[suffix_prefix] = subroi_dict
        elif 'rectum_dilated_upper_half' == para_name:
            # 直肠上部分  且是整个最小矩形区域都是mask 而不只是直肠区域
            slice_mask2 = generate_dilated_upper_half_mask(slice_mask, dilation_iter=my, propotion=mx)  # propotion是只保留上百分比的区域   dilation_iter是外扩距离 无内收参数
            # 特殊ID 靠近施源器部分 给赋值0
            slice_mask2 = process_ID_of_rectum_mask(date_day, ID, z, add_para, slice_mask2)
            sub_img = slice_img.copy()
            subroi_dict = {}
            for tag in mask_tags:
                if tag != 'rectum_dilated_upper_half':
                    raise ValueError('tag=rectum_dilated_upper_half 参数目前需要统一格式')
                sub_img_d, mask_d = extract_roi_crop(ID, date_day, z, sub_img, slice_mask2)  # 该函数将大mask 和subimg 缩小为最小外接矩形的形式
                # 特殊处理 为了计算整个最小矩形区域的指标  将mask缩减为最小矩形后再赋值为全0
                mask_d[:] = 1
                save_subroi_to_ct_results(ct_results, folder_name, base_roi,
                                          suffix_prefix, tag, z, sub_img_d, mask_d)
                subroi_dict[tag] = {'img': sub_img_d.astype(np.float32),
                                    'mask': mask_d.astype(np.uint8)}
            mask_dict[suffix_prefix] = subroi_dict
        elif 'rectum_ring_LCR' == para_name:  # 解包的参数后三个是没有意义的
            masks_dict = generate_ring_mask(mask_tags, slice_mask, dilation_iter=my, erosion_iter=mx)
            sub_img = slice_img.copy()
            subroi_dict = {}
            for tag in mask_tags:
                if tag not in masks_dict:
                    continue
                m = masks_dict[tag]
                # 特殊ID 靠近施源器部分 给赋值0
                m = process_ID_of_rectum_mask(date_day, ID, z, add_para, m)
                sub_img_d, mask_d = extract_roi_crop(ID, date_day, z, sub_img, m)  # 该函数将大mask 和subimg 缩小为最小外接矩形的形式
                save_subroi_to_ct_results(ct_results, folder_name, base_roi,
                                          suffix_prefix, tag, z, sub_img_d, mask_d)
                subroi_dict[tag] = {'img': sub_img_d.astype(np.float32),
                                    'mask': mask_d.astype(np.uint8)}
            mask_dict[suffix_prefix] = subroi_dict
    return mask_dict


def metric_mtf(context):  # col 代表该指标是反映竖直边缘的性能  row是水平边缘
    img = context['img']
    spacing = context['spacing']
    roi = context['roi']
    ID = context['ID']
    date = context['date']
    z = context['z']
    mtf_areas_shuzhi, cutoff_freqs_shuzhi, mtf_areas_shuiping, cutoff_freqs_shuiping = compute_mtf_both_axes(roi, ID, date, z, img, float(
        spacing[0]))
    mtf_areas_shuzhi_old, mtf_areas_shuiping_old = mtf_calcu(img)  # 这里反过来解包  因为 计算行 代表的是竖直边缘的mtf
    return {
        'mtf_areas_shuzhi': mtf_areas_shuzhi,
        'cutoff_freqs_shuzhi': cutoff_freqs_shuzhi,
        'mtf_areas_shuiping': mtf_areas_shuiping,
        'cutoff_freqs_shuiping': cutoff_freqs_shuiping,  # cutoff_freq 越高通常意味着图像系统的分辨率越好。
        'oldmtf_areas_shuzhi': mtf_areas_shuzhi_old,
        'oldmtf_areas_shuiping': mtf_areas_shuiping_old,
    }


# def metric_nmtf(context):
#     img = context['img']
#     nimg = (img - img.min()) / (img.ptp() + 1e-8)
#     ctx_dict = dict(context, img=nimg)
#     return {f"n_{k}": v for k, v in metric_mtf(ctx_dict).items()}
def metric_nmtf(context):
    img = context['img']
    ptp = np.nanmax(img) - np.nanmin(img)
    if img.size == 0 or np.all(np.isnan(img)) or ptp < 1e-8:
        nimg = img
    else:
        nimg = (img - np.nanmin(img)) / (ptp + 1e-8)
    ctx_dict = dict(context, img=nimg)
    return {f"n_{k}": v for k, v in metric_mtf(ctx_dict).items()}


def metric_zmtf(context):
    img = context['img']
    if img.size == 0 or np.all(np.isnan(img)) or np.nanstd(img) < 1e-8:
        zimg = img  # 或 return 固定结果
    else:
        zimg = (img - np.nanmean(img)) / (np.nanstd(img) + 1e-8)
    ctx_dict = dict(context, img=zimg)
    return {f"z_{k}": v for k, v in metric_mtf(ctx_dict).items()}


# def metric_zmtf(context):
#     img = context['img']
#     zimg = (img - img.mean()) / (img.std() + 1e-8)
#     ctx_dict = dict(context, img=zimg)
#     return {f"z_{k}": v for k, v in metric_mtf(ctx_dict).items()}
def metric_radius(context, th_list=(3500, 3000, 2500, 2000, 1500, 1000, 500, 300)):
    img = context['img']
    spacing = context['spacing']
    return {f"radius_{th}": round(cal_radius(img, spacing, th), 2) for th in th_list}


def roi_allowed(metric_limit_dict, metric_name: str, roi_name: str) -> bool:
    """若 metric 没有限制或 roi 在白名单内，返回 True"""
    allowed = metric_limit_dict.get(metric_name)
    return (allowed is None) or (roi_name in allowed)


def diameter_cal_pre_process(slice_img, slice_mask, y_min_roi, y_max_roi, x_min_roi, x_max_roi, ):
    # 扩大周围区域 用来统计不同阈值下中心行列的像素个数 mask不变 依然是施源器的mask  y_min_roi, y_max_roi, x_min_roi, x_max_roi,代表原始mask最大范围，margin参数代表根据最大范围外扩图像和mask
    sub_img_diameter, slice_mask_diameter = crop_subregion(slice_img, slice_mask, y_min_roi,
                                                           y_max_roi, x_min_roi, x_max_roi,
                                                           margin_y_min=10, margin_x_min=10,
                                                           margin_y_max=10, margin_x_max=10)
    # ▶️ 提前检查是否为3个连通区域
    labeled, n_cc = ndi.label(slice_mask_diameter)
    # 计算连通区域的质心，提取左右两管
    cc_props = ndi.center_of_mass(slice_mask_diameter, labeled, range(1, n_cc + 1))
    cc_sorted = sorted(enumerate(cc_props, start=1), key=lambda x: x[1][1])  # 按X坐标排序
    labels_to_use = [cc_sorted[0][0], cc_sorted[2][0]]  # 最左和最右两个label
    return labeled, labels_to_use, sub_img_diameter, slice_mask_diameter, n_cc


def metric_diameter(context):
    """
    只对 spc1 的 ROI 做 diameter 指标，返回一个 {指标名: 数值} 的字典
    context: dict, 包含 slice_img, slice_mask, y_min_roi, ..., ID, z, roi 等
    """
    if context['roi'] != 'spc1':
        return {}
    labeled, labels_to_use, sub_img_d, slice_mask_d, n_cc = diameter_cal_pre_process(
        context['slice_img'], context['slice_mask'],
        context['y_min_roi'], context['y_max_roi'],
        context['x_min_roi'], context['x_max_roi'],
    )
    if n_cc != 3:
        return {}  # 非 3 管结构，跳过
    res = {}
    for th in [5000, 4000, 3000, 2000, 1500, 1000, 500, 300, 100, 50, 0]:
        d_row, d_col = cal_diameter(
            context['ID'], context['date'], context['z'], context['roi'],
            sub_img=sub_img_d,
            slice_mask=slice_mask_d,
            spacing=context['spacing'],
            th=th,
            labeled=labeled,
            labels_to_use=labels_to_use,
        )
        res[f'row_d_{th}'] = round(d_row, 2)
        res[f'col_d_{th}'] = round(d_col, 2)
    return res


def extract_features(image_sitk, mask_sitk):
    """
    使用pyradiomics提取特征
    """
    # 初始化特征提取器，并指定只计算第一阶统计特征中的均匀性
    extractor = featureextractor.RadiomicsFeatureExtractor()
    extractor.disableAllFeatures()
    extractor.enableFeatureClassByName('firstorder')
    # 启用特定的一阶特征：均匀性、熵、偏度、峰度
    extractor.enableFeaturesByName(firstorder=['Uniformity', 'Entropy', 'Skewness', 'Kurtosis'])
    features = extractor.execute(image_sitk, mask_sitk)
    return features


def manual_texture_features(img, mask):
    """
    手动计算 texture features（支持任意小尺寸 mask）
    支持维度：[H,W] 或 [D,H,W]
    """
    values = img[mask > 0].flatten()
    values = values.astype(np.float32)

    if values.size == 0:
        return {
            'Uniformity': 0,
            'Entropy': 0,
            'Skewness': 0,
            'Kurtosis': 0
        }
    # 计算概率直方图（可调 bins）
    hist, _ = np.histogram(values, bins=64, density=True)
    prob = hist[hist > 0]
    entropys_num = -np.sum(prob * np.log2(prob))
    uniformity_num = np.sum(prob ** 2)
    skewness_num = skew(values)
    kurtosis_num = kurtosis(values)
    return {
        'Uniformity': uniformity_num,
        'Entropy': entropys_num,
        'Skewness': skewness_num,
        'Kurtosis': kurtosis_num
    }


def calculate_total_average_gradient(pixel_matrix, mask_matrix):
    # 将输入转换为NumPy数组
    pixel_matrix = np.array(pixel_matrix, dtype=np.float32)
    mask_matrix = np.array(mask_matrix, dtype=np.uint8)
    # 现在 矩阵是仅裁减了矩形的  mask则是进一步用于裁剪roi的
    #  先计算所有梯度 再裁剪roi区域
    # 计算水平梯度
    sobelx1 = cv2.Sobel(pixel_matrix, cv2.CV_64F, 1, 0, ksize=3)
    # 计算竖直方向梯度
    sobelx2 = cv2.Sobel(pixel_matrix, cv2.CV_64F, 0, 1, ksize=3)
    # 再次裁剪去掉最外面一层  因为边缘梯度是0 会影响
    masked_gradients1 = sobelx1[1:-1, 1:-1].copy()
    masked_gradients2 = sobelx2[1:-1, 1:-1].copy()
    masked_mask_matrix = mask_matrix[1:-1, 1:-1].copy()
    # 将掩码应用于梯度矩阵，只保留roi内的值
    # roi_masked_gradients1 = cv2.bitwise_and(masked_gradients1, masked_gradients1, mask=masked_mask_matrix)
    # roi_masked_gradients2 = cv2.bitwise_and(masked_gradients2, masked_gradients2, mask=masked_mask_matrix)
    roi_masked_gradients1 = masked_gradients1 * (masked_mask_matrix > 0)
    roi_masked_gradients2 = masked_gradients2 * (masked_mask_matrix > 0)

    # # 提取ROI内的非零元素
    # 经验证 切片操作有效
    # 接下来我们计算梯度矩阵的平均梯度 和标准差梯度
    total_average_gradient = mean_gra(roi_masked_gradients1, roi_masked_gradients2, masked_mask_matrix)
    total_average_std_gradient = std_gra(roi_masked_gradients1, roi_masked_gradients2)
    return total_average_gradient, total_average_std_gradient


def mean_gra(masked_gradients1, masked_gradients2, mask_matrix):
    # 计算掩码区域内所有梯度值的总和
    total_gradient_sum1 = np.sum(masked_gradients1)
    total_gradient_sum2 = np.sum(masked_gradients2)
    # 计算掩码区域内的非零像素总数
    total_nonzero_pixels = np.count_nonzero(mask_matrix)
    # 计算x方向总的平均梯度
    total_average_gradient1 = total_gradient_sum1 / total_nonzero_pixels if total_nonzero_pixels > 0 else 0
    # 计算y方向总的平均梯度
    total_average_gradient2 = total_gradient_sum2 / total_nonzero_pixels if total_nonzero_pixels > 0 else 0
    # 总平均梯度
    total_average_gradient = (total_average_gradient2 + total_average_gradient1) / 2
    return total_average_gradient


def std_gra(masked_gradients1, masked_gradients2):
    # 提取非零区域，或非全 0 区域来计算稳定 std
    valid_g1 = masked_gradients1[masked_gradients1 != 0]
    valid_g2 = masked_gradients2[masked_gradients2 != 0]
    if valid_g1.size < 2 or np.allclose(valid_g1, valid_g1[0]):
        std1 = -999.0
    else:
        std1 = np.std(valid_g1)
    if valid_g2.size < 2 or np.allclose(valid_g2, valid_g2[0]):
        std2 = -999.0
    else:
        std2 = np.std(valid_g2)
    return (std1 + std2) / 2


def calculate_blur(image):
    image = np.uint8(image * 255)
    # 应用Laplacian算子并计算方差作为模糊度量
    laplacian = cv2.Laplacian(image, cv2.CV_64F)
    variance = laplacian.var()
    return variance


def calculate_edge_blur(image, mask):
    image = np.uint8(image * 255)
    distances = []  # 存储每一行的25%到75%的像素数量
    for y in range(image.shape[0]):
        # 获取当前行及对应行的掩码
        row = image[y, :]
        row_mask = mask[y, :]
        # 获取掩码内的像素值
        row_pixels = row[row_mask > 0]
        # 如果当前行没有有效像素，跳过
        if len(row_pixels) == 0:
            continue
        # 计算当前行的25%和75%分位数
        lower_quantile = np.percentile(row_pixels, 10)
        upper_quantile = np.percentile(row_pixels, 90)
        # 找出每一行中符合25%到75%范围的像素数量
        valid_pixels = np.logical_and(row >= lower_quantile, row <= upper_quantile)
        distance = np.sum(valid_pixels)  # 计算符合条件的像素数量
        distances.append(distance)
    # 计算平均距离(mm)，假设每个像素间距1mm
    average_distance = np.mean(distances)
    return average_distance


def generate_range_pairs(range1_list, range2_list):
    """
    根据给定的多个 range1 和 range2 百分位范围，生成所有组合的参数对。

    参数：
        range1_list: list of tuple，如 [(0, 5), (5, 10)]
        range2_list: list of tuple，如 [(90, 95), (95, 100)]

    返回：
        list of dict，每个 dict 可直接用于调用 add_percentile_mean_diff
    """
    pairs = []
    for r1 in range1_list:
        for r2 in range2_list:
            key_suffix = f'{r1[0]}{r1[1]}{r2[0]}{r2[1]}'
            pairs.append({
                'range1': r1,
                'range2': r2,
                'key_suffix': key_suffix
            })
    return pairs


def add_percentile_diff(result, pixels, range1=(0, 5), range2=(95, 100), key_suffix=''):
    """
    计算指定两个百分位区间内像素均值的差，并写入 result 字典。

    参数：
        result : dict，结果字典
        pixels : 1D numpy array，像素值数组
        range1 : tuple，百分位范围1，例如 (0, 5)
        range2 : tuple，百分位范围2，例如 (95, 100)
        key    : str，结果中对应的键名
    """
    p1_min, p1_max = np.percentile(pixels, range1)
    p2_min, p2_max = np.percentile(pixels, range2)
    group1 = pixels[(pixels >= p1_min) & (pixels <= p1_max)]
    group2 = pixels[(pixels >= p2_min) & (pixels <= p2_max)]
    mean1 = np.mean(group1) if group1.size > 0 else np.nan
    mean2 = np.mean(group2) if group2.size > 0 else np.nan
    result[f'meandiff_{key_suffix}'] = -float(mean2 - mean1)  # 负数 让相关性反过来
    result[f'stddiff_{key_suffix}'] = -float(np.std(pixels))


def get_percentile_slice(pixels, low_percent, high_percent):
    pixels_sorted = np.sort(pixels)
    n = len(pixels_sorted)
    idx_low = int(np.floor(n * low_percent / 100))
    idx_high = int(np.ceil(n * high_percent / 100))
    return pixels_sorted[idx_low:idx_high]


def add_percentile_diff_special(result, pixels, pixels_special, range1=(80, 100), range2=(0, 20), key_suffix=''):
    group1 = get_percentile_slice(pixels, *range1)
    group2 = get_percentile_slice(pixels_special, *range2)
    mean1 = np.mean(group1)
    mean2 = np.mean(group2)
    median1 = np.median(group1)
    median2 = np.median(group2)
    result[f'ring_outinner_meandiff_{key_suffix}'] = -float(abs(mean2 - mean1))
    result[f'ring_outinner_meddiff_{key_suffix}'] = -float(abs(median2 - median1))
    result[f'ring_outinner_meanratio_{key_suffix}'] = -float(abs(mean2 / mean1))
    result[f'ring_outinner_medratio_{key_suffix}'] = -float(abs(median2 / median1))


def add_threshold_ratio_features(result, pixels_all, up_percentiles, low_percentiles, percentiles):
    # 先计算百分位数指标
    for p in percentiles:
        key2 = f'percentile_{p}'
        result[key2] = float(np.percentile(pixels_all, p))
    # === Step 2: 将百分位数转为实际像素值阈值
    up_thresholds = {p: np.percentile(pixels_all, p) for p in up_percentiles}
    low_thresholds = {p: np.percentile(pixels_all, p) for p in low_percentiles}
    for t1, t2 in itertools.product(up_thresholds, low_thresholds):
        pos_count = np.sum(pixels_all > t1)
        neg_count = np.sum(pixels_all < t2)
        pos_vals = pixels_all[pixels_all > t1]
        neg_vals = pixels_all[pixels_all < t2]
        ratio_key = f'count_up{t1}low{t2}_ratio'
        diff_key = f'value_up{t1}low{t2}_diff'
        mean_ratio_key = f'value_up{t1}low{t2}_ratio'
        neg_mean = np.mean(neg_vals) if neg_vals.size > 0 else np.nan
        pos_mean = np.mean(pos_vals) if pos_vals.size > 0 else np.nan
        if neg_count == 0:
            result[ratio_key] = -999
        else:
            result[ratio_key] = pos_count / neg_count
        result[diff_key] = abs(pos_mean - neg_mean) if neg_vals.size > 0 else -999
        # 安全设置：避免除以 0 或 nan
        if np.isnan(neg_mean) or neg_mean == 0:
            result[mean_ratio_key] = -999
        else:
            result[mean_ratio_key] = abs(pos_mean / neg_mean)
        # # 同时计算和percentile的组合权重（后面会补percentile值）
        # for p in percentiles:
        #     perc_key = f'percentile_{p}'
        #     combo_key_1 = f'combo_{ratio_key}_{perc_key}_w_half'
        #     combo_key_2 = f'combo_{ratio_key}_{perc_key}_w_13_23'
        #     combo_key_3 = f'combo_{ratio_key}_{perc_key}_w_23_13'
        #     if perc_key in result:
        #         result[combo_key_1] = 0.5 * result[ratio_key] + 0.5 * result[perc_key]
        #         result[combo_key_2] = (1 / 3) * result[ratio_key] + (2 / 3) * result[perc_key]
        #         result[combo_key_3] = (2 / 3) * result[ratio_key] + (1 / 3) * result[perc_key]


def metric_static(context):
    sub_img = context['img']
    #  注意由于对于派生和原始roi都是用一个mask文本  所以派生的时候需要替换掉原始context的mask
    mask = context['mask']
    result = {}
    pixels = sub_img[mask == 1]
    pixels_special = sub_img[mask == 2]  # 这个是针对某些特殊的mask 中 2的位置 比如内环1 外环2 这种mask 用来计算特殊指标
    pixels_all = sub_img[mask > 0]
    # tag = context['tag']
    if pixels.size < 2 or np.all(pixels == pixels[0]):
        result.update(dict.fromkeys([
            'std', 'range', 'mean', 'median',
            'gra_num', 'grastd_num', 'blur_num', 'edge_num'
        ], -999.0))
        result.update(manual_texture_features(sub_img, mask))
        return result
    # gra_num, grastd_num = calculate_total_average_gradient(sub_img, mask)
    # blur_num = calculate_blur(sub_img)
    # edge_num = calculate_edge_blur(sub_img, mask)
    # === 添加指标 ===
    result['std'] = -float(np.std(pixels))  # 标准差（可选）
    result['range'] = float(np.max(pixels) - np.min(pixels))  # 范围
    result['mean'] = -float(np.mean(pixels))
    result['median'] = -float(np.median(pixels))
    # 求阈值分割（比如0HU） 正负像素数量比值  以及  其他百分位数指标
    add_threshold_ratio_features(result, pixels_all, [50, 60, 70, 80, 85, 95], [5, 10, 15, 25, 35, 45],
                                 [95, 94, 93, 92, 91, 90, 89, 88, 87, 86, 85, 84, 83, 82, 81, 80, 79, 78, 77, 76, 75, 74, 73, 72, 71, 70, 69, 68, 67, 66, 65, 64, 63, 62, 61, 60])
    # 求低于某一阈值下的所有像素的均值 等
    # percentiles = [5, 10, 25, 50, 75]  # 你可以按需设置
    # for p in percentiles:
    #     threshold = np.percentile(pixels, p)
    #     below_pixels = pixels[pixels < threshold]
    #     result[f'mean_below_p{p}'] = float(np.mean(below_pixels)) if below_pixels.size > 0 else -999
    #     result[f'med_below_p{p}'] = float(np.median(below_pixels)) if below_pixels.size > 0 else -999
    # result['gra_num'] = float(gra_num)
    # result['grastd_num'] = float(grastd_num)
    # result['blur_num'] = float(blur_num)
    # result['edge_num'] = float(edge_num)

    # result['plus_minus_count0_bizhi'] = np.sum(pixels_all > 0) / np.sum(pixels_all < 0)
    # 指标测试：统计特定HU范围内像素数量和百分比
    # add_hu_range_counts(result, pixels, hu_ranges = generate_hu_ranges([-150,-125,-100,-75,-50,-25,-60,-40,-30,-15], [-22,25,0,-15,-30,-45,-60]))
    # 指标测试： 对前多少百分位的像素统一计算统计指标
    # add_percentile_stats(result, pixels, [2,4,6,8,10,13,16,19,22,25,28,31,34,37,40,43,46,49,52,55,58,61,64,67,70])
    # 指标测试：对rectum_ring_LCR左右两侧计算更多指标
    # if 'rectum_ring_LCR' in context['derived_roi']:
    #     # # 计算环状三个roi的 百分比均值差和标准差差
    #     # range1s = [(0, 5), (5, 10), (10, 15), (15, 20),(20, 25)]
    #     # range2s = [(80, 90), (90, 100), (70, 80), (65, 70),(60, 65), (55, 60), (50, 55), (45, 50)]
    #     # pairs = generate_range_pairs(range1s, range2s)
    #     # for pair in pairs:
    #     #     add_percentile_diff(result, pixels,range1=pair['range1'], range2=pair['range2'],key_suffix=pair['key_suffix'])
    #     # 计算特殊的内外环的均值差指标
    #     range1_special = [(0, 10), (10, 20),(20,30), (30, 40),(40,50), (50, 60),(60,70), (70,80), (80,90), (90,100)]
    #     range2_special =  [(0, 10), (10, 20),(20,30), (30, 40),(40,50), (50, 60),(60,70), (70,80), (80,90), (90,100)]
    #     pair_specials = generate_range_pairs(range1_special, range2_special)
    #     for pair_special in pair_specials:
    #         add_percentile_diff_special(result, pixels, pixels_special, range1=pair_special['range1'], range2=pair_special['range2'], key_suffix=pair_special['key_suffix'])
    # 指标rectum_ring_l  rectum_ring_r
    # if 'rectum' == context['roi']: # 需要重新对slice_img进行处理 生成roi  否则外部遍历是一个roi遍历 没法算
    #     slice_img = context['slice_img']
    #     slice_mask = context['slice_mask']
    #     diff_left = compute_side_statdiff(slice_img, slice_mask, side='l', outer_inner_pairs=[(1,1), (2,2), (3,3)], stats=['mean', 'median'])
    #     diff_right = compute_side_statdiff(slice_img, slice_mask, side='r', outer_inner_pairs=[(1,1), (2,2), (3,3)], stats=['mean', 'median'])
    #     for key in diff_left.keys():
    #         result[f'rectum_{key}'] = -diff_left[key]  # 注意取负号
    #         result[f'rectum_{key.replace("l_", "r_")}'] = -diff_right[key.replace("l_", "r_")]
    #         # 平均值与最大值
    #         mean_val = float(diff_left[key] + diff_right[key.replace("l_", "r_")]) / 2
    #         max_val = max(diff_left[key], diff_right[key.replace("l_", "r_")])
    #         result[f'rectum_{key.replace("l_", "meanlr_")}'] = mean_val
    #         result[f'rectum_{key.replace("l_", "maxlr_")}'] = max_val
    # === 添加影像组学类指标 ===
    # img_sitk = sitk.GetImageFromArray(sub_img.astype(np.float32))
    # mask_sitk = sitk.GetImageFromArray(mask.astype(np.uint8))
    # if min(sub_img.shape) <= 1 or min(mask.shape) <= 1:
    #     result.update(manual_texture_features(sub_img, mask))
    # else:
    #     features = extract_features(img_sitk, mask_sitk)
    #     result['Entropy'] = float(features.get('original_firstorder_Entropy'))  # 信息熵
    #     result['Uniformity'] = float(features.get('original_firstorder_Uniformity'))  # 均匀性
    #     result['Skewness'] = float(features.get('original_firstorder_Skewness'))
    #     result['Kurtosis'] = float(features.get('original_firstorder_Kurtosis'))
    return result


def compute_side_statdiff(slice_img, slice_mask, side='l', outer_inner_pairs=[(1, 1), (2, 2)], stats=['mean', 'median']):
    """
    根据不同膨胀腐蚀参数，计算图像在指定mask边缘的多个统计指标差异（左右两侧可选）

    参数：
        slice_img: 图像切片，2D数组
        slice_mask: 掩膜，2D 0/1 数组
        side: 'left' 或 'right'
        outer_inner_pairs: [(d_outer, d_inner), ...]
        stats: ['mean', 'median', ...]

    返回：
        results: 字典，包含各参数组合和统计方式的差异结果
    """
    results = {}
    mask = (slice_mask > 0).astype(np.uint8)
    cols = np.any(mask, axis=0)
    x_min, x_max = np.where(cols)[0][[0, -1]]
    x_mid = (x_min + x_max) // 2
    if side == 'l':
        side_mask = np.zeros_like(mask)
        side_mask[:, x_min:x_mid + 1] = mask[:, x_min:x_mid + 1]
    elif side == 'r':
        side_mask = np.zeros_like(mask)
        side_mask[:, x_mid + 1:x_max + 1] = mask[:, x_mid + 1:x_max + 1]
    else:
        raise ValueError("side must be 'left' or 'right'")
    for d_outer, d_inner in outer_inner_pairs:
        dilated = cv2.dilate(side_mask, None, iterations=d_outer)
        eroded = cv2.erode(side_mask, None, iterations=d_inner)
        outer_ring = (dilated - side_mask).astype(bool)
        inner_ring = (side_mask - eroded).astype(bool)
        outer_vals = slice_img[outer_ring]
        inner_vals = slice_img[inner_ring]
        if outer_vals.size == 0 or inner_vals.size == 0:
            for stat in stats:
                key = f'{side}_d{d_outer}{d_inner}{stat}'
                results[key] = np.nan
            continue
        for stat in stats:
            if stat == 'mean':
                val = abs(np.mean(outer_vals) - np.mean(inner_vals))
            elif stat == 'median':
                val = abs(np.median(outer_vals) - np.median(inner_vals))
            else:
                continue  # 可扩展更多统计方式
            key = f'{side}_d{d_outer}{d_inner}{stat}'
            results[key] = float(val)
    return results


def generate_hu_ranges(lower_bounds, upper_bounds):
    """
    从下界列表和上界列表生成合法的 HU 范围组合 (lower, upper)，要求 lower < upper。
    """
    ranges = []
    for low in lower_bounds:
        for high in upper_bounds:
            if low < high:
                ranges.append((low, high))
    return ranges


def add_hu_range_counts(result, pixels, hu_ranges):
    for hu_min, hu_max in hu_ranges:
        key = f'count_percent_{int(hu_min)}_{int(hu_max)}'
        count = np.sum((pixels >= hu_min) & (pixels <= hu_max)) / pixels.size
        result[key] = float(count)
        key2 = f'count_{int(hu_min)}_{int(hu_max)}'
        count2 = np.sum((pixels >= hu_min) & (pixels <= hu_max))
        result[key2] = float(count2)


def add_percentile_stats(result, pixels, percentiles: list):
    """
    为给定多个百分位数阈值，计算像素中对应区间的统计指标，并更新到 result 字典中。
    """
    pixels = np.asarray(pixels).flatten()
    for p in percentiles:
        sub_pixels = pixels[pixels <= np.percentile(pixels, p)]
        if sub_pixels.size < 2:
            result[f'low_{p}_mean'] = np.nan
            result[f'low_{p}_std'] = np.nan
            result[f'low_{p}_range'] = np.nan
            result[f'low_{p}_median'] = np.nan
        else:
            result[f'low_{p}_mean'] = float(np.mean(sub_pixels))
            result[f'low_{p}_std'] = float(np.std(sub_pixels))
            result[f'low_{p}_range'] = float(np.max(sub_pixels) - np.min(sub_pixels))
            result[f'low_{p}_median'] = float(np.median(sub_pixels))


def metric_nstatic(context):
    """
    计算 min-max 归一化图像的统计指标
    """
    img = context['img']
    ptp = img.ptp()
    if ptp < 1e-8:
        norm_img = img
    else:
        norm_img = (img - img.min()) / (ptp + 1e-8)
    ctx_d = dict(context, img=norm_img)
    return {f"n_{k}": v for k, v in metric_static(ctx_d).items()}


def metric_zstatic(context):
    """
    计算 z-score 标准化图像的统计指标
    """
    img = context['img']
    std = img.std()
    if std < 1e-8:
        std_img = np.zeros_like(img)
    else:
        std_img = (img - img.mean()) / (std + 1e-8)
    ctx_d = dict(context, img=std_img)
    return {f"z_{k}": v for k, v in metric_static(ctx_d).items()}


def metric_nps(ctx):
    """
    计算 NPS 五个基础指标：
    nps_low / nps_high / nps_ring / nps_total / nps_directional
    所需字段：
        ctx['img']       → 计算图像
        ctx['spacing']   → 像素间距
        ctx['tag'] → 'all' | 'up' | 'down' | 'left' | 'right'
    """
    img = ctx['img']
    spacing = ctx['spacing']
    tag = ctx['tag']
    # # roi = ctx['roi']
    # # ID = ctx['ID']
    # # date = ctx['date']
    # folder = ctx['folder']
    total_power_spectrum, total_nps_ring, total_directional_power, _, low_freq_nps, high_freq_nps = calculate_nps(img, spacing, tag=tag)
    return {
        # 'nps_low': low_freq_nps,
        # 'nps_high': high_freq_nps,
        'nps_ring': total_nps_ring,
        # 'nps_total': total_power_spectrum,
        # 'nps_directional': total_directional_power,
    }


def combine_nps_metric(context, region, suffix):
    # ① 检查 tags 是否齐
    required = {'up', 'left', 'right'}
    if not required.issubset(region.keys()):
        missing = required - region.keys()
        raise ValueError(f"spc1_for_MAS 缺少方向 {missing} (suffix={suffix})")
    # ➤ ② 三个方向各算一次 ―― 用通用 metric_nps
    # 三方向分别跑 metric_nps。metric_nps 要读 ctx['tag'](或 direction)，统一用 tag：
    mk_ctx = lambda d: dict(context,
                            img=region[d]['img'],
                            mask=None,  # NPS 不用 mask
                            tag=d)  # 全程统一用 tag
    res_up, res_left, res_right = map(metric_nps, map(mk_ctx, ('up', 'left', 'right')))
    # ③ 组合公式
    # nps_high_rl_num = (res_left['nps_high'] + res_right['nps_high']) / 2
    # nps_low_rl_num = (res_left['nps_low'] + res_right['nps_low']) / 2
    # nps_low_url_num = (res_up['nps_low'] + res_left['nps_low'] + res_right['nps_low']) / 3
    nps_ring_rl_num = (res_left['nps_ring'] + res_right['nps_ring']) / 2
    # nps_total_rl_num = (res_left['nps_total'] + res_right['nps_total']) / 2
    # nps_directional_rl_num = (res_left['nps_directional'] + res_right['nps_directional']) / 2
    # nps_low_rl_and_direct1 = (nps_low_rl_num + nps_directional_rl_num) / 2
    # nps_low_rl_and_direct2 = (nps_directional_rl_num * 3 + nps_low_rl_num) / 4
    # nps_low_rl_and_direct3 = (nps_directional_rl_num + nps_low_rl_num * 3) / 4
    # return nps_high_rl_num,nps_low_rl_num,nps_low_url_num,nps_ring_rl_num, nps_total_rl_num, nps_directional_rl_num, nps_low_rl_and_direct1, nps_low_rl_and_direct2, nps_low_rl_and_direct3
    return nps_ring_rl_num


def cal_noref_metric_for_single_folder_ID_date_roi_metric(idx, para_mask_tag_map, parameters, ct_folder, structure_name, structure_control, metrics, if_save_cal_image, img_save_folder, if_isolate_img_folder):
    # print(f"[子进程 {idx}] 正在处理: {ct_folder}")
    ct_results = OrderedDict()
    # print(f'读取 {ct_folder}')
    folder_name = os.path.basename(ct_folder)
    # 目前的区分文件夹之间的标签  以后根据目的可以改动
    add_tag = os.path.basename(ct_folder).split('_')[-1]
    # 获取 RS 文件路径
    rs_file = glob.glob(os.path.join(ct_folder, "RS*.dcm"))[0]  # 取第一个匹配的 RS 文件
    # 加载 CT 图像
    # print(f'获取CT图像相关参数')
    volume, origin, spacing, num_slices, height, width, dcm_slices, date_day, date, ID = load_dicom_images(ct_folder)
    # print(ID,date)
    folder_dict = ct_results.setdefault(folder_name, OrderedDict())
    folder_dict['ID'] = ID
    folder_dict['date'] = date
    # 对每个存在的ROI进行处理
    # if isinstance(structure_name, str):
    #     structure_name = [structure_name]
    # if idx == 0:
    #     print(f'     子进程 {idx} 共{len(structure_name)}个roi')
    for roi in structure_name:
        # print(f'       当前遍历 {folder_name}  -  {roi}')
        mask_3d = load_structure_mask(rs_file, volume.shape, dcm_slices, structure_name=roi)
        mask_3d_spc1 = load_structure_mask(rs_file, volume.shape, dcm_slices, structure_name='spc1')  # 提前缓存一个 某些roi派生后续要用
        if structure_control:
            mask_3d_control = np.zeros_like(volume, dtype=np.uint8)
            layer_roi_map = {z: [] for z in range(volume.shape[0])}  # 初始化计算每层指标时的控制结构记录
            for roi_control in structure_control:
                try:
                    mask_3d_sub_control = load_structure_mask(rs_file, volume.shape, dcm_slices,
                                                              structure_name=roi_control)
                    # 记录哪些结构在每层中出现
                    for z in range(mask_3d_sub_control.shape[0]):
                        if np.any(mask_3d_sub_control[z]):  # 如果该层有掩膜
                            layer_roi_map[z].append(roi_control)
                    mask_3d_control = np.logical_or(mask_3d_control, mask_3d_sub_control).astype(np.uint8)
                except ValueError as e:
                    # print(f"⚠️ 跳过结构 '{roi_control}'：{e}")
                    pass
        else:
            layer_roi_map = {z: [] for z in range(volume.shape[0])}
            mask_3d_control = mask_3d
        # 如果mask_3d为空跳过本次roi循环
        if not mask_3d.any():
            continue
        # 初始化 roi 结果字典
        ct_results[folder_name].setdefault(roi, {})
        # 对每个切片进行处理
        for z in range(num_slices):
            # 获取 z 坐标值
            z_pos = dcm_slices[z].ImagePositionPatient[2]
            # print(f'           计算第 {z}  层')
            # 获取该层的掩膜 + 控制掩膜因为计算的rmse指标有些层面太靠近spc1的硬化区域而较高但却不反映对比度的值。
            slice_mask = mask_3d[z]
            slice_mask_control = mask_3d_control[z]
            if np.any(slice_mask) and np.any(slice_mask_control):  # true代表该层面有限定范围的辅助结构
                # 记录该层面控制结构参与情况（如 'rib_heart'，若无则为 'none'）
                control_structures = layer_roi_map.get(z, [])
                control_label = '_'.join(control_structures) if control_structures else 'none'
                ct_results[folder_name][roi].setdefault('roi_control', {})[z] = control_label
                # ✅ 记录该层面的坐标
                ct_results[folder_name][roi].setdefault('zcoord', {})[z] = z_pos / 10
                # 1、常规处理roi
                # 先统一获取层面图像和roi范围
                slice_img, y_min_roi, y_max_roi, x_min_roi, x_max_roi = get_slice_mask(z, volume, slice_mask)
                # 0 裁剪最小外切矩形 +  1 常规进一步裁剪图像
                sub_img, slice_mask2 = crop_subregion(slice_img, slice_mask, y_min_roi, y_max_roi, x_min_roi, x_max_roi, )
                # 除上面的特殊处理外， 对于其他正常结构 为计算全参考指标 对每一个z层面都保存了准备用于计算的图像
                ct_results[folder_name][roi].setdefault('sub_images', {})[z] = sub_img.astype(np.float32)
                # 2、 对于需要派生或各种特殊处理的roi 单独进行 不要覆盖正常roi遍历
                mask_dict_all = OrderedDict()
                if roi == 'spc1' and 'spc1_for_MAS' in parameters:  # 对于spc1  特殊生成spc1周围的 roi的mask  # 判断参数中是否有针对该ROI的参数键  有的话则根据参数进一步生成基于该roi的子roi
                    mask_dict_all['spc1_for_MAS'] = gen_subroi_masks(mask_3d, ID, date_day, z, volume,
                                                                     ct_results, folder_name, roi, 'spc1_for_MAS', parameters,
                                                                     slice_img=slice_img, slice_mask=slice_mask,
                                                                     y_min_roi=y_min_roi, y_max_roi=y_max_roi, x_min_roi=x_min_roi, x_max_roi=x_max_roi,
                                                                     mask_tags=para_mask_tag_map['spc1_for_MAS']
                                                                     )
                if roi == 'spc1' and 'spc1_for_LCR' in parameters:
                    mask_dict_all['spc1_for_LCR'] = gen_subroi_masks(mask_3d, ID, date_day, z, volume,
                                                                     ct_results, folder_name, roi, 'spc1_for_LCR', parameters,
                                                                     slice_img=slice_img, slice_mask=slice_mask,
                                                                     y_min_roi=y_min_roi, y_max_roi=y_max_roi, x_min_roi=x_min_roi, x_max_roi=x_max_roi,
                                                                     mask_tags=para_mask_tag_map['spc1_for_LCR']
                                                                     )
                if roi == 'HRCTV' and 'HRCTV_for_LCR' in parameters:
                    mask_dict_all['HRCTV_for_LCR'] = gen_subroi_masks(mask_3d, ID, date_day, z, volume,
                                                                      ct_results, folder_name, roi, 'HRCTV_for_LCR', parameters,
                                                                      slice_img=slice_img, slice_mask=slice_mask,
                                                                      y_min_roi=y_min_roi, y_max_roi=y_max_roi, x_min_roi=x_min_roi, x_max_roi=x_max_roi,
                                                                      mask_tags=para_mask_tag_map['HRCTV_for_LCR']
                                                                      )
                if roi == 'rectum' and 'rectum_for_LCR' in parameters:
                    mask_dict_all['rectum_for_LCR'] = gen_subroi_masks(mask_3d, ID, date_day, z, volume,
                                                                       ct_results, folder_name, roi, 'rectum_for_LCR', parameters,
                                                                       slice_img=slice_img, slice_mask=slice_mask,
                                                                       y_min_roi=y_min_roi, y_max_roi=y_max_roi, x_min_roi=x_min_roi, x_max_roi=x_max_roi,
                                                                       mask_tags=para_mask_tag_map['rectum_for_LCR']
                                                                       )
                if roi == 'rectum' and 'rectum_dilated_upper_half' in parameters:
                    mask_dict_all['rectum_dilated_upper_half'] = gen_subroi_masks(mask_3d, ID, date_day, z, volume,
                                                                                  ct_results, folder_name, roi, 'rectum_dilated_upper_half', parameters,
                                                                                  slice_img=slice_img, slice_mask=slice_mask,
                                                                                  y_min_roi=y_min_roi, y_max_roi=y_max_roi, x_min_roi=x_min_roi, x_max_roi=x_max_roi,
                                                                                  mask_tags=para_mask_tag_map['rectum_dilated_upper_half'], add_para=mask_3d_spc1  # 这里不用管 就是这样
                                                                                  )
                if roi == 'rectum' and 'rectum_ring_LCR' in parameters:
                    mask_dict_all['rectum_ring_LCR'] = gen_subroi_masks(mask_3d, ID, date_day, z, volume,
                                                                        ct_results, folder_name, roi, 'rectum_ring_LCR',
                                                                        parameters,
                                                                        slice_img=slice_img, slice_mask=slice_mask,
                                                                        y_min_roi=y_min_roi, y_max_roi=y_max_roi,
                                                                        x_min_roi=x_min_roi, x_max_roi=x_max_roi,
                                                                        mask_tags=para_mask_tag_map['rectum_ring_LCR'], add_para=mask_3d_spc1)

                # if idx == 0:
                #     print(f"        🧷 {idx}进程 打印{roi}的 mask_dict_all 结构")
                #     summarize_mask_dict_structure(mask_dict_all, max_depth=10)
                # 3先保存用于真实计算的图像区域 和修改每个计算层面中指定ROI的范围的特殊处理
                sub_img_save, slice_mask_save = crop_subregion(slice_img, slice_mask, y_min_roi, y_max_roi, x_min_roi, x_max_roi, margin_y_min=70, margin_x_min=70, margin_y_max=70, margin_x_max=70)
                # 4 定义指标计算的纳入和排除映射表
                # 这个是你要计算的指标及其对应的函数
                METRIC_FN = {
                    'mtf': metric_mtf,
                    'nmtf': metric_nmtf,
                    'zmtf': metric_zmtf,
                    'radius': metric_radius,
                    'diameter': metric_diameter,
                    'static': metric_static,
                    'nstatic': metric_nstatic,
                    'zstatic': metric_zstatic,
                    'nps': metric_nps,
                }
                # 若 metric 没有限制或 roi 在白名单列表内，则会在原始roi和派生roi中计算
                metric_limit_dict = {
                    'radius': ['spc1'],  # 例如 这就会限制 radius 指标的 roi 只能是 spc1 不能是其他原始roi 也不是任何派生roi
                    'diameter': ['spc1'],
                }
                # 统一调用的输入参数字典
                context = {
                    'folder': folder_name,
                    'img': sub_img,
                    'mask': slice_mask2,
                    'slice_mask': slice_mask,
                    'spacing': spacing,
                    'roi': roi,
                    'ID': ID,
                    'date': date,
                    'z': z,
                    'roi_control': control_label,
                    'zcoord': z_pos / 10,
                    'slice_img': slice_img,
                    'y_min_roi': y_min_roi,
                    'y_max_roi': y_max_roi,
                    'x_min_roi': x_min_roi,
                    'x_max_roi': x_max_roi,
                    'tag': 'all',
                    'derived_roi': ''
                }
                # assert isinstance(sub_img, np.ndarray), f"metric_mtf: 'img' {roi}{ID}{date}{z}不是 ndarray，而是 {type(sub_img)}"
                # 5 开始计算指标
                if idx == 0:
                    print(f'           {idx}  一共{len(metrics)}个指标')
                for metric in metrics:
                    # ---------- 原始 ROI的相关指标（除static  nps）计算 ----------
                    if metric in METRIC_FN:
                        # 对原始roi计算指标
                        if roi_allowed(metric_limit_dict, metric, roi):  # 可以过滤metric_limit_dict中你确定不需要计算的指标
                            result_dict = METRIC_FN[metric](context)
                            if not result_dict:
                                continue  # 跳过空结果
                            for metric_name, v in result_dict.items():
                                write_metric(ct_results, folder_name, roi, metric_name, z, v)

                        # ---------- 派生 ROI ----------
                        count = 0  # ✅ 初始化计数器
                        for para_key, one_mask_dict in mask_dict_all.items():
                            tags = para_mask_tag_map.get(para_key)
                            if tags is None:
                                raise ValueError(f"⚠️ 未知派生 ROI 可能未提前定义para_mask_tag_map中的tags: {para_key}")

                            # 对每个派生roi计算指标
                            for suffix, region in one_mask_dict.items():
                                # suffix形如 不同参数组合成的  suffix_prefix =f"{my}{mx}{My}{Mx}{pixels_far}"
                                for tag in tags:  # 对于派生roi  变化了是img和mask 所以替换掉输入的参数
                                    img = region[tag]['img']
                                    mask = region[tag]['mask']
                                    # 派生roi名字要和前面一致
                                    derived_roi = f"{roi}_{suffix}_{tag}"
                                    if not roi_allowed(metric_limit_dict, metric, derived_roi):
                                        continue
                                    # -------- 派生 ROI 初始化（只做一次） --------
                                    dr_dict = ct_results[folder_name].setdefault(derived_roi, OrderedDict())
                                    dr_dict.setdefault('roi_control', {})[z] = control_label
                                    dr_dict.setdefault('zcoord', {})[z] = z_pos / 10
                                    # ctx_d是重点 区别于原始roi  遍历派生roi时 需要替换初始化的context参数 对不同指标函数替换都要考虑到
                                    # 例如 对于原始ROI是 slice_mask2   对于派生roi就是对应的mask
                                    # 对于其他参数如x_max_roi 可以更新替换但目前用不着 没有派生情况下的函数使用它
                                    ctx_d = dict(context, img=img, mask=mask, tag=tag, derived_roi=derived_roi)
                                    # 执行统一计算
                                    result_dict = METRIC_FN[metric](ctx_d)
                                    if not result_dict:
                                        continue
                                    for k, v in result_dict.items():
                                        write_metric(ct_results, folder_name, derived_roi, k, z, v)
                                    count += 1  # ✅ 每次成功写入后计数
                            # ---------- 特殊指标如nps组合 左右取最小等  仅在tag循环之后  重计算 ----------
                            if roi == 'spc1' and metric == 'nps' and (para_key in ['rectum_for_LCR', 'spc1_for_MAS', 'HRCTV_for_LCR', 'spc1_for_LCR', '']):
                                for suffix, region in one_mask_dict.items():
                                    # nps_high_rl_num, nps_low_rl_num, nps_low_url_num, nps_ring_rl_num, nps_total_rl_num, nps_directional_rl_num, nps_low_rl_and_direct1, nps_low_rl_and_direct2, nps_low_rl_and_direct3= combine_nps_metric(context, region, suffix)
                                    nps_ring_rl_num = combine_nps_metric(context, region, suffix)

                                    W = lambda name, val: write_metric(ct_results, folder_name, roi, f'{name}_{suffix}', z, val)
                                    # W('nps_high_rl', nps_high_rl_num)
                                    # W('nps_low_rl', nps_low_rl_num)
                                    # W('nps_low_url', nps_low_url_num)
                                    W('nps_ring_rl', nps_ring_rl_num)
                                    # W('nps_total_rl', nps_total_rl_num)
                                    # W('nps_directional_rl', nps_directional_rl_num)
                                    # W('nps_low_rl_and_direct1', nps_low_rl_and_direct1)
                                    # W('nps_low_rl_and_direct2', nps_low_rl_and_direct2)
                                    # W('nps_low_rl_and_direct3', nps_low_rl_and_direct3)
                                    count += 1  # ✅ 每次成功写入后计数
                        # if idx == 0:
                        #     print(f"              ✅ { idx}进程  总共遍历{roi}的派生 ROI 并写入 {metric} 指标的次数为: {count}")
                # 2 最后再保存本次计算的文件夹的计算层面图像
                if if_save_cal_image == 1 and roi == 'rectum':
                    # 如果文件夹不存在，则创建
                    os.makedirs(img_save_folder, exist_ok=True)
                    #  保存图像
                    # 图像显示参数
                    # vmin, vmax = 40 - 350 / 2, 40 + 350 / 2  # CBCT 窗宽窗位  -200 - 1600 / 2, -200 + 1600 / 2
                    # 腹部为
                    vmin, vmax = -125, 225
                    if if_isolate_img_folder == 1:
                        save_single_image(slice_mask_save, sub_img_save, os.path.join(img_save_folder, os.path.basename(ct_folder),
                                                                                      f'{ID}_{date_day}_{roi}_{z}.jpg'), vmin, vmax)
                    else:
                        if add_tag != date_day:
                            save_single_image(slice_mask_save, sub_img_save,
                                              os.path.join(img_save_folder, f'{add_tag}_{ID}_{date_day}_{roi}_{z}.jpg'),
                                              vmin, vmax)
                        else:
                            save_single_image(slice_mask_save, sub_img_save,
                                              os.path.join(img_save_folder, f'{ID}_{date_day}_{roi}_{z}.jpg'), vmin,
                                              vmax)
        # print(f'       {folder_name}-{roi}-结束')
    # print(f'                       当前遍历文件夹 {folder_name} 结束')
    return ct_results


def process_multiple_ct_folders(group_idx, para_mask_tag_map, parameters, ref_tag, if_isolate_img_folder, if_save_cal_image, out_dir, ct_folders, structure_name, structure_control=None, metrics=None):
    """
    处理多个 CT 文件夹，计算多个图像指标（如 MTF、标准差、噪声功率谱等）。
    """
    img_save_folder = os.path.join(out_dir, '1计算层面图像')
    os.makedirs(out_dir, exist_ok=True)
    os.makedirs(img_save_folder, exist_ok=True)
    if metrics is None:
        # 默认计算 MTF 和噪声功率谱，如果需要添加其他指标，可以扩展这个列表
        # metrics = ['mtf', 'std', 'nps']
        raise ValueError('必须指定你要计算的指标')
    # 不并行计算的版本
    # for ct_folder in ct_folders:
    #     ct_results = cal_noref_metric_for_single_folder_ID_date_roi_metric(ct_folder, structure_name, structure_control,
    #                                                            metrics, if_save_cal_image, img_save_folder,
    #                                                            if_isolate_img_folder)
    # 并行计算
    print("▶️ 开始并行计算 ...")
    start_all = time.time()
    start_parallel = time.time()
    # args_list = [(para_mask_tag_map,parameters,ct_folder, structure_name, structure_control,
    #         metrics, if_save_cal_image, img_save_folder,
    #         if_isolate_img_folder)for ct_folder in ct_folders]
    args_list = [(i, para_mask_tag_map, parameters, ct_folder, structure_name, structure_control,
                  metrics, if_save_cal_image, img_save_folder, if_isolate_img_folder)
                 for i, ct_folder in enumerate(ct_folders)]
    # 并行运行每个文件夹的无参考指标计算
    with Pool(processes=cpu_count()) as pool:
        results = pool.starmap(cal_noref_metric_for_single_folder_ID_date_roi_metric, args_list)
    print(f'第{group_idx}组文件夹  并行计算结束 用时 {time.time() - start_parallel:.2f} 秒 开始合并指标')
    # 合并每个子字典  这里是遍历每个子进程计算得到的字典结构   每个子进程计算的folder name必须是唯一  否则后面的进程结果会替换前面的
    start_merge = time.time()
    ct_results = OrderedDict()
    for r in results:
        ct_results.update(r)
    print(f"✅ ")
    total_entries = 0
    for folder_result in ct_results.values():
        for roi_result in folder_result.values():
            total_entries += len(roi_result)  # 每个 roi_result 是 z -> metric 的字典
    print(f"✅           第{group_idx}组文件夹  合并无参考指标结果用时 {time.time() - start_merge:.2f} 秒  共 {total_entries} z层面的指标结果被合并")
    # 全参考指标
    start_ref = time.time()
    compute_multi_reference_metrics_grouped(
        ct_results,
        roi_list=structure_name,  # 可以是一个或多个 ROI  这里没有计算派生的roi
        ref_keywords=ref_tag,  # 用于识别参考图像的关键词（从 Folder 最后一个段落中提取）
        metrics=metrics  # 要计算的全参考指标
    )
    print(f'第{group_idx}组文件夹  全参考指标计算结束  {time.time() - start_ref:.2f} 秒   总耗时: {time.time() - start_all:.2f} 秒  开始保存结果')
    return ct_results, img_save_folder


def save_ct_results_to_excel(group_idx,
                             ct_results,
                             out_dir_new,
                             all_metric_data_file_name,
                             all_metric_avg_data_file_name
                             ):
    ori_rows = []
    start_read = time.time()
    for folder, fdata in ct_results.items():
        fid = fdata.get('ID')
        fdate = fdata.get('date')
        for roi, rdata in fdata.items():
            if roi in ('ID', 'date'):
                continue
            z_coord_map = rdata.get('zcoord', {})
            roi_control_map = rdata.get('roi_control', {})
            for metric, mdata in rdata.items():
                if metric in ('zcoord', 'roi_control', 'sub_images'):  # ✅ 增加 sub_images 排除
                    continue
                for z, ref_dict in mdata.items():
                    if not isinstance(ref_dict, dict):
                        raise TypeError(
                            f"[错误] {folder}→{roi}→{metric}→{z} 的值应为 dict，但实际为：{type(ref_dict)}"
                        )
                    for ref_img, record in ref_dict.items():
                        z_coord = z_coord_map.get(z)
                        roi_control = roi_control_map.get(z)
                        tag = record.get('add_tag', '')  # 视情况添加  如果前面赋值结果的函数没有使用add_tag之类的 tag  那这里也不会报错
                        ori_rows.append([
                            folder, fid, fdate, ref_img,
                            roi, metric, z, z_coord,
                            roi_control, record['value'],
                            tag
                        ])
    # 写入 DataFrame
    ori_df = pd.DataFrame(ori_rows, columns=["Folder", "ID", "Date", "ref_image", "ROI", "Metric", "Z", "Z_Coord", "roi_control", "Value", "add_tag"])
    print(f"✅ 将结果转换为df 准备保存。   用时: {time.time() - start_read:.2f} 秒")
    start_process = time.time()
    # 删除其中metric = sub_images的行
    ori_df = ori_df[ori_df['Metric'] != 'sub_images']
    # 强制将 Value 转换为 float，非数字（如数组、列表）将变成 NaN
    ori_df['Value'] = pd.to_numeric(ori_df['Value'], errors='raise')
    # Value 四舍五入
    ori_df['Value'] = ori_df['Value'].round(3)
    # 生成每组内部的 Z_normal
    # print(ori_df)
    ori_df['Z_normal'] = ori_df.groupby(["Folder", "ID", "Date", 'ref_image', 'ROI', 'Metric', 'roi_control'])['Z'].rank(method='first').astype(int)
    ori_df["Group"] = ori_df.groupby(["Folder", "ID", "Date", 'ref_image', ]).ngroup() + 1  # 日期作为分组依据
    ori_df = ori_df.sort_values(by=['ROI', 'Metric', "Folder", 'ref_image', 'Z'])
    # 将 'Value' 列移动到最后一列
    value_col = ori_df.pop('Value')  # 先弹出
    ori_df['Value'] = value_col  # 再添加回去到最后一列
    # 每次计算可能有一些特殊分类用的tag  比如算法 所以目前默认新增一个tag列 通用folder最后一个元素
    ori_df['add_tag'] = ori_df['Folder'].str.split('_').str[-1]
    # print('可选   本次特殊处理 按照每个add_tag列中9种不同算法的组   将组内数据归一化添加一列  ')
    ori_df = normalize_metric_values(ori_df, group_keys=['ID', "Date", "ref_image", "ROI", "Metric", 'roi_control', 'Z'], method='z-score', new_col='group_norm_Value', )  # z-score   min-max
    # 保存原始数据
    os.makedirs(out_dir_new, exist_ok=True)
    all_metric_data_file_path = os.path.join(out_dir_new, f'{all_metric_data_file_name}-{group_idx}.csv')
    print(f"✅ 转换df后的预处理用时: {time.time() - start_process:.2f} 秒")
    start_save = time.time()
    # 保存为csv
    ori_df.to_csv(all_metric_data_file_path, index=False)
    # 计算均值表  默认有 add_tag 列  可以算进去求均值用 不过folder列已经包含该信息了  主要就是最小分组的求不同层面的均值
    df_averaged = (ori_df
                   .groupby(["Folder", "ID", "Date", "ref_image", "ROI", "Metric", 'roi_control', 'add_tag'], as_index=False)[["Value", "group_norm_Value"]]
                   .mean().rename(columns={"Value": "AvgValue", "group_norm_Value": "Avggroup_norm_Value"}))
    df_averaged = df_averaged.sort_values(by=['ROI', 'Metric', "Folder", 'ref_image'])
    df_averaged["Group"] = df_averaged.groupby(["Folder", "ID", "Date", 'ref_image', ]).ngroup() + 1
    all_metric_avg_data_file_path = os.path.join(out_dir_new, f'{all_metric_avg_data_file_name}-{group_idx}.csv')
    df_averaged.to_csv(all_metric_avg_data_file_path, index=False)
    print(f"✅ 数据ori_df  df_averaged 已保存  用时:  {time.time() - start_save:.2f} 秒 ")


def merge_excel_results(out_dir, output_dir_new, prefix='data', avg_prefix='dataAvg'):
    # print(output_dir)
    # for f in glob.glob(os.path.join(output_dir, f"{prefix}*.csv")):
    #     print("✅ 匹配到文件：", f)
    #     print("🧪 提取数字：", re.findall(rf"{prefix}-(\d+)\.csv", os.path.basename(f)))
    # 获取所有分组保存的结果文件 并按照数字排序
    data_files = sorted(
        glob.glob(os.path.join(output_dir_new, f"{prefix}-*.csv")),
        key=lambda x: int(re.findall(rf"{prefix}-(\d+)\.csv", os.path.basename(x))[0])
    )
    avg_files = sorted(
        glob.glob(os.path.join(output_dir_new, f"{avg_prefix}-*.csv")),
        key=lambda x: int(re.findall(rf"{avg_prefix}-(\d+)\.csv", os.path.basename(x))[0])
    )
    # print( f"✅ 获取所有分组保存的结果文件 并按照数字排序：{data_files}  {avg_files}")
    # 合并所有 data 表格
    df_all = pd.concat([pd.read_csv(f) for f in data_files], ignore_index=True)
    df_avg_all = pd.concat([pd.read_csv(f) for f in avg_files], ignore_index=True)
    # 保存为统一命名
    df_all.to_csv(os.path.join(out_dir, f"{prefix}.csv"), index=False)
    df_avg_all.to_csv(os.path.join(out_dir, f"{avg_prefix}.csv"), index=False)
    print(f"✅ 合并完成：{prefix}.csv 和 {avg_prefix}.csv")


def group_ct_folders_by_id_only(ct_folder_path_all, id_per_group=5):
    ct_folders_all = [os.path.join(ct_folder_path_all, folder) for folder in os.listdir(ct_folder_path_all)
                      if os.path.isdir(os.path.join(ct_folder_path_all, folder))]
    # ✅ 提取 ID（不含日期）
    id_groups = defaultdict(list)
    for folder in ct_folders_all:
        folder_name = os.path.basename(folder)
        ID = folder_name.split("_")[0]  # ← 只取 ID 部分
        id_groups[ID].append(folder)
    all_ids = list(id_groups.keys())
    grouped_folders = []
    for i in range(0, len(all_ids), id_per_group):
        group_ids = all_ids[i:i + id_per_group]
        group = []
        for gid in group_ids:
            group.extend(id_groups[gid])
        grouped_folders.append(group)
    return grouped_folders


# 本代码计算图像的指标  ct_folder_path_all 包含若干个子文件夹，每个文件夹包含一套CT RS  文件夹名应当是ID_date_tag
# 保存每次计算的层面图像  # 如果你以后用于计算常规指标 把 其中mask_spc1的内容删掉就行
# 默认并行计算多个文件夹  每个文件夹内 可以计算多参数生成的ROI  需要自定义修改
if __name__ == "__main__":
    ct_folder_path_all = r"D:\std\科研项目数据汇总\999后装施源器伪影研究+SPC\0数据\7-1SPC监测伪影和勾画-患者图像"
    out_dir = r"D:\ob仓库\研究\2工作\临床与科研工作\999后装施源器伪影研究+SPC\1分析\7-1SPC监测伪影和勾画\0指标筛选"
    # 想计算的指标  可多个   'diameter','nps','ssim','psnr','ncc','rmse','radius','mtf','zmtf','nmtf'  'static'  'nstatic'    'zstatic'
    metric_list = ['nps']
    # 目标roi 可以算所有文件中只要存在的roi
    ROI = ['spc1']  # ,'rectum' 'HRCTV' 'spc1',
    ref_text = ['ori']  # 计算全参考指标用的参考文件夹名字中的biaoqian/tag  指定你标识参考图像的标识符  即你的图像文件夹命名应该是ID+date+标识符tag 给全参考指标用  可以多个  但只会在同一个ID-date组下计算  为空则计算所有两两图像比较
    # 控制计算层面的roi  如果不为空 则计算无参考指标时计算和它共存的层面  'HRCTV'  或者None  多个则会累加只要存在的层面
    roi_control = ['HRCTV']  # 'zc','zc2','zc3'
    if_save_cal_image = 1  # 保存图像是否  如果需要保存指定roi 在函数中修改
    if_isolate_img_folder = 11  # 保存图像是否进一步单独建文件夹
    # 对大文件夹按ID分组 指定每组ID数量 用于限定每次并行的文件夹数量 注意是ID 每个ID可能包含很多分次
    id_per_group = 200
    # 用于派生不同roi的img外扩参数， 可能包含1或多个计算范围参数组的嵌套列表

    # # 手动给定
    # parameters_single = {'spc1_for_MAS':[[9, 3, 9, 3, 1, 'param_150']]}   #  'spc1_for_MAS':[[9, 3, 9, 3, 1, 'param_150']]
    # # 自动生成：{'spc1':generate_parameter_groups_for_spc1(y_range=(2, 10), y_step=1, x_range=(2, 10), x_step=1,inner_range=(1, 5))}
    # parameters = {
    #     # 'spc1_for_MAS':generate_parameter_groups_for(y_range=(7, 11), y_step=1, x_range=(2, 5), x_step=1,inner_range=(0, 3)),
    #     # 'spc1_for_LCR':generate_parameter_groups_for(y_range=(2, 10), y_step=1, x_range=(2, 10), x_step=1,inner_range=(1, 3)),
    #     # 'HRCTV_for_LCR': generate_parameter_groups_for(y_range=(2, 10), y_step=1, x_range=(2, 10), x_step=1,inner_range=(-2, 3)),
    #     # 'rectum_for_LCR': generate_parameter_groups_for(y_range=(2, 6), y_step=6, x_range=(2, 6), x_step=6,inner_range=(-3, 3)),  # 负数代表向内裁剪
    #     # 'rectum_dilated_upper_half':generate_two_param_sets(range(1,10), [1/3,2/3,1]),  # 上半不同大小的直肠区域 依次代表外扩范围和 保留上百分比区域
    #     # 'rectum_ring_LCR':generate_two_param_sets(range(1, 3), range(5,12)),  # 环状派生roi 依次代表 外扩和内缩 范围
    #     # 'rectum_dilated_upper_half': generate_two_param_sets(range(3,6), [1 / 4, 1 / 3, 1 / 2, 2 / 3]),
    #               }
    # # 在给定外扩参数下的mask-tag 即不同的派生方式 常规有上下左右四种派生 可以添加其他派生方式 在函数中定义
    # para_mask_tag_map = {
    #     'HRCTV_for_LCR': ('up', 'down', 'left', 'right'),
    #     'spc1_for_LCR': ('up', 'down', 'left', 'right'),
    #     'rectum_for_LCR': ('up', 'left', 'right'),
    #     'rectum_dilated_upper_half': ('rectum_dilated_upper_half',), # 注意必要有逗号
    #     'spc1_for_MAS': ('up', 'down', 'left', 'right'),
    #     'rectum_ring_LCR': ('rectum_ring_LCR',),# 注意必要有逗号  可以附加新的参数 ：'rectum_ring_l','rectum_ring_r','ring_include_boundry'
    # }
    # # 开始计算
    # if parameters_single:
    #     parameters = parameters_single
    # pd.set_option('display.max_rows', 4000)  # 最多显示 1000 行
    # pd.set_option('display.max_columns', 100)  # 最多显示 100 列
    # pd.set_option('display.width', 200)  # 控制总宽度，防止自动换行
    # pd.set_option('display.max_colwidth', None)  # 每列最大字符长度，None 表示不限制
    # # ct_folders = [os.path.join(ct_folder_path_all, folder) for folder in os.listdir(ct_folder_path_all)
    # #               if os.path.isdir(os.path.join(ct_folder_path_all, folder))]  # 只选择文件夹

    grouped_folders = group_ct_folders_by_id_only(ct_folder_path_all, id_per_group=id_per_group)
    # 打印parameters中各参数组的大小
    for key, value in parameters.items():
        print(f"本次计算的 {key} 参数组大小：{len(value)}")
    # 按组计算
    out_dir_new = os.path.join(out_dir, f'0原始计算结果')
    os.makedirs(out_dir_new, exist_ok=True)
    # 先清空out_dir_new中的所有文件
    for file_name in os.listdir(out_dir_new):
        os.remove(os.path.join(out_dir_new, file_name))
    for group_idx, ct_folders in enumerate(grouped_folders):
        group_idx += 1
        print(f"\n🚀 正在处理第 {group_idx}/{len(grouped_folders)} 组，共 {len(ct_folders)} 个文件夹...")
        ct_results, img_save_folder = process_multiple_ct_folders(group_idx, para_mask_tag_map, parameters, ref_text, if_isolate_img_folder, if_save_cal_image, out_dir, ct_folders, ROI, structure_control=roi_control, metrics=metric_list)
        save_ct_results_to_excel(group_idx, ct_results, out_dir_new, f'data', f'dataAvg')
    merge_excel_results(out_dir, out_dir_new, prefix='data', avg_prefix='dataAvg')
